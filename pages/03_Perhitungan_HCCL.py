import streamlit as st

# Cek apakah sudah login dari halaman utama
if "login_status" not in st.session_state or not st.session_state["login_status"]:
    st.error("🚨 Akses Ditolak! Silakan login di halaman utama terlebih dahulu.")
    st.stop()

from style_utils import apply_custom_style
apply_custom_style()

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime

# ============================
# KONFIGURASI GLOBAL CL
# ============================
COL_RMCC = 'CONCENTRATION LIMIT USULAN RMCC'
COL_LISTED = 'CONCENTRATION LIMIT TERKENA % LISTED SHARES'
COL_FF = 'CONCENTRATION LIMIT TERKENA % FREE FLOAT'
COL_PERHITUNGAN = 'CONCENTRATION LIMIT SESUAI PERHITUNGAN'
THRESHOLD_5M = 5_000_000_000
KODE_EFEK_KHUSUS = ['LPKR', 'MLPL', 'NOBU', 'PTPP', 'SILO']
OVERRIDE_MAPPING = {
    'LPKR': 10_000_000_000, 'MLPL': 10_000_000_000,
    'NOBU': 10_000_000_000, 'PTPP': 50_000_000_000, 'SILO': 10_000_000_000
}
TARGET_100 = 100.0
TOLERANCE = 1e-6

# ===============================================================
# FUNGSI UTILITAS UNTUK CONCENTRATION LIMIT (CL)
# ===============================================================

def calc_concentration_limit_listed(row):
    try:
        if row['PERBANDINGAN DENGAN LISTED SHARES (Sesuai Perhitungan)'] >= 0.05:
            return 0.0499 * row['LISTED SHARES'] * row['CLOSING PRICE']
        return None
    except Exception:
        return None

def calc_concentration_limit_ff(row):
    try:
        if row['PERBANDINGAN DENGAN FREE FLOAT (Sesuai Perhitungan)'] >= 0.20:
            return 0.1999 * row['FREE FLOAT (DALAM LEMBAR)'] * row['CLOSING PRICE']
        return None
    except Exception:
        return None

def override_rmcc_limit(row):
    kode = row['KODE EFEK']
    nilai_rmcc = row[COL_RMCC]
    if kode in OVERRIDE_MAPPING:
        if nilai_rmcc == 0.0:
            return 0.0
        return min(nilai_rmcc, OVERRIDE_MAPPING[kode])
    return nilai_rmcc

def reset_concentration_limit(
    df_main: pd.DataFrame,
    haircut_col: str = "HAIRCUT PEI USULAN DIVISI",
    conc_limit_col: str = "CONCENTRATION LIMIT USULAN RMCC",
    conc_calc_col: str = "CONCENTRATION LIMIT SESUAI PERHITUNGAN",
    tolerance: float = 1e-6,
    threshold_limit: float = 5_000_000_000,
    inplace: bool = True
) -> pd.DataFrame:
    if not inplace:
        df_main = df_main.copy()

    if haircut_col not in df_main.columns:
        df_main[haircut_col] = 0.0

    df_main[haircut_col] = pd.to_numeric(df_main[haircut_col], errors='coerce')
    df_main[conc_calc_col] = pd.to_numeric(df_main[conc_calc_col], errors='coerce')

    valid_haircut = df_main[haircut_col].dropna()
    target_100_reset = 100.0 if not valid_haircut.empty and valid_haircut.max() > 1 + tolerance else 1.0

    mask_haircut_100 = (df_main[haircut_col].sub(target_100_reset).abs() < tolerance)
    mask_below_threshold = (df_main[conc_calc_col] < threshold_limit)
    mask_final = mask_haircut_100 | mask_below_threshold

    df_main.loc[mask_final, conc_limit_col] = 0.0
    return df_main

def keterangan_uma(uma_date):
    if pd.notna(uma_date):
        if not isinstance(uma_date, datetime):
            try:
                uma_date = pd.to_datetime(str(uma_date))
            except Exception:
                return "Sesuai Metode Perhitungan"
        return f"Sesuai Haircut KPEI, mempertimbangkan pengumuman UMA dari BEI tanggal {uma_date.strftime('%d %b %Y')}"
    return "Sesuai Metode Perhitungan"

# ===============================================================
# FUNGSI UTAMA PERHITUNGAN CONCENTRATION LIMIT
# ===============================================================

def calculate_concentration_limit(df_cl_source: pd.DataFrame) -> pd.DataFrame:
    df = df_cl_source.copy()

    if 'KODE EFEK' not in df.columns:
        df = df.rename(columns={df.columns[0]: 'KODE EFEK'})

    df['KODE EFEK'] = df['KODE EFEK'].astype(str).str.strip()

    # 1. Limit marjin baru
    df['SAHAM MARJIN BARU?'] = df['SAHAM MARJIN BARU?'].astype(str).str.upper().str.strip()
    df['CONCENTRATION LIMIT KARENA SAHAM MARJIN BARU'] = np.where(
        df['SAHAM MARJIN BARU?'] == 'YA',
        df[COL_PERHITUNGAN] * 0.50,
        df[COL_PERHITUNGAN]
    )

    # 2. Hitung limit listed & FF
    df[COL_LISTED] = df.apply(calc_concentration_limit_listed, axis=1)
    df[COL_FF] = df.apply(calc_concentration_limit_ff, axis=1)

    # 3. Ambil nilai minimum
    limit_cols_for_min = [
        'CONCENTRATION LIMIT KARENA SAHAM MARJIN BARU',
        COL_LISTED, COL_FF, COL_PERHITUNGAN
    ]
    df['MIN_CL_OPTION'] = df[limit_cols_for_min].fillna(np.inf).min(axis=1)

    # 4. Tentukan pemicu nol
    mask_pemicu_nol = (
        (df['CONCENTRATION LIMIT KARENA SAHAM MARJIN BARU'].fillna(np.inf) < THRESHOLD_5M) |
        (df[COL_PERHITUNGAN].fillna(np.inf) < THRESHOLD_5M) |
        (df[COL_LISTED].fillna(np.inf) < THRESHOLD_5M) |
        (df[COL_FF].fillna(np.inf) < THRESHOLD_5M)
    )

    df[COL_RMCC] = np.where(mask_pemicu_nol, 0.0, df['MIN_CL_OPTION'])

    # 5. Override emiten khusus (CAP 10M/50M)
    mask_not_zero = (df[COL_RMCC] != 0.0)
    df.loc[mask_not_zero, COL_RMCC] = df.loc[mask_not_zero].apply(override_rmcc_limit, axis=1).round(0)

    # 6. Haircut usulan divisi
    df['HAIRCUT KPEI'] = pd.to_numeric(df['HAIRCUT KPEI'], errors='coerce').fillna(0)
    df['HAIRCUT PEI USULAN DIVISI'] = np.where(
        (df['UMA'].fillna('-') != '-') & pd.notna(df['UMA']),
        df['HAIRCUT KPEI'],
        df['HAIRCUT PEI']
    )

    # 7. Nolkan CL jika haircut 100% atau CL perhitungan < 5M
    df = reset_concentration_limit(df)

    # 7B. Set haircut jadi 100% jika CL = 0
    HAIRCUT_COL_USULAN = 'HAIRCUT PEI USULAN DIVISI'
    mask_rmcc_nol = (df[COL_RMCC] == 0)
    df['TEMP_HAIRCUT_VAL_ASLI'] = pd.to_numeric(df[HAIRCUT_COL_USULAN], errors='coerce')
    df.loc[mask_rmcc_nol, HAIRCUT_COL_USULAN] = 1.0

    # 8. Keterangan awal
    df['PERTIMBANGAN DIVISI (HAIRCUT)'] = df['UMA'].apply(keterangan_uma)
    df['PERTIMBANGAN DIVISI (CONC LIMIT)'] = 'Sesuai metode perhitungan'

    # 9. Definisikan masker untuk keterangan
    mask_emiten = df['KODE EFEK'].isin(KODE_EFEK_KHUSUS)
    mask_nol_final = (df[COL_RMCC] == 0) & (~mask_emiten)

    mask_lt5m_strict = (
        (df['CONCENTRATION LIMIT KARENA SAHAM MARJIN BARU'].fillna(np.inf) < THRESHOLD_5M) |
        (df[COL_PERHITUNGAN].fillna(np.inf) < THRESHOLD_5M) |
        (df[COL_LISTED].fillna(np.inf) < THRESHOLD_5M) |
        (df[COL_FF].fillna(np.inf) < THRESHOLD_5M) |
        (df[COL_RMCC].fillna(np.inf) < THRESHOLD_5M)
    ) & mask_nol_final

    mask_hc100_origin = (
        (df['TEMP_HAIRCUT_VAL_ASLI'].sub(1.0).abs() < TOLERANCE) |
        (df['TEMP_HAIRCUT_VAL_ASLI'].sub(TARGET_100).abs() < TOLERANCE)
    ) & mask_nol_final & (~mask_lt5m_strict)

    mask_non_nol = (df[COL_RMCC] != 0)
    df['CL_OVERRIDE_VAL'] = df['KODE EFEK'].map(OVERRIDE_MAPPING).fillna(np.inf)
    mask_emiten_override = mask_non_nol & mask_emiten & \
                           (df[COL_RMCC].round(0) == df['CL_OVERRIDE_VAL'].round(0))
    mask_other_non_nol = mask_non_nol & (~mask_emiten_override)

    mask_kriteria_listed = pd.notna(df[COL_LISTED])
    mask_kriteria_ff = pd.notna(df[COL_FF])

    mask_listed_saja = mask_other_non_nol & \
                       (df[COL_RMCC].round(0) == df[COL_LISTED].round(0))

    mask_ff_saja = mask_other_non_nol & \
                   (df[COL_RMCC].round(0) == df[COL_FF].round(0)) & \
                   (~mask_listed_saja)

    mask_marjin_baru = mask_other_non_nol & \
                       (df[COL_RMCC].round(0) == df['CONCENTRATION LIMIT KARENA SAHAM MARJIN BARU'].round(0)) & \
                       (df['SAHAM MARJIN BARU?'] == 'YA') & \
                       (~mask_listed_saja) & (~mask_ff_saja)

    mask_listed_ff_upgrade = (mask_listed_saja | mask_ff_saja) & \
                              mask_kriteria_listed & mask_kriteria_ff

    # 10. Penerapan keterangan (prioritas rendah ke tinggi)
    df.loc[mask_lt5m_strict, 'PERTIMBANGAN DIVISI (CONC LIMIT)'] = 'Penyesuaian karena Batas Konsentrasi < Rp5 Miliar'
    df.loc[mask_lt5m_strict, 'PERTIMBANGAN DIVISI (HAIRCUT)'] = 'Penyesuaian karena Batas Konsentrasi 0'
    df.loc[mask_hc100_origin, 'PERTIMBANGAN DIVISI (CONC LIMIT)'] = 'Penyesuaian karena Haircut PEI 100%'
    df.loc[mask_emiten & mask_nol_final, 'PERTIMBANGAN DIVISI (CONC LIMIT)'] = 'Penyesuaian karena profil emiten'
    df.loc[mask_emiten_override, 'PERTIMBANGAN DIVISI (CONC LIMIT)'] = 'Penyesuaian karena profil emiten'
    df.loc[mask_marjin_baru, 'PERTIMBANGAN DIVISI (CONC LIMIT)'] = 'Penyesuaian karena saham baru masuk marjin'
    df.loc[mask_listed_saja, 'PERTIMBANGAN DIVISI (CONC LIMIT)'] = 'Penyesuaian karena melebihi 5% listed shares'
    df.loc[mask_ff_saja, 'PERTIMBANGAN DIVISI (CONC LIMIT)'] = 'Penyesuaian karena melebihi 20% free float'
    df.loc[mask_listed_ff_upgrade, 'PERTIMBANGAN DIVISI (CONC LIMIT)'] = 'Penyesuaian karena melebihi 5% listed & 20% free float'

    # Cleanup kolom bantuan
    df = df.drop(columns=['MIN_CL_OPTION', 'TEMP_HAIRCUT_VAL_ASLI', 'CL_OVERRIDE_VAL'], errors='ignore')

    return df

# ===============================================================
# FUNGSI INJECTOR KE TEMPLATE EXCEL (dari Code 1, disesuaikan)
# ===============================================================

def update_excel_template(file_template, df_hasil):
    """
    Menyuntikkan hasil perhitungan ke template Excel.
    
    Sheet CONC (data mulai row 5):
      col 16 (P) = CONCENTRATION LIMIT KARENA SAHAM MARJIN BARU
      col 17 (Q) = CONCENTRATION LIMIT TERKENA % LISTED SHARES
      col 18 (R) = CONCENTRATION LIMIT TERKENA % FREE FLOAT
      col 19 (S) = CONCENTRATION LIMIT USULAN RMCC
      col 22 (V) = PERTIMBANGAN DIVISI (CONC LIMIT)
    
    Sheet HC (data mulai row 5):
      col 18 (R) = HAIRCUT PEI USULAN DIVISI
      col 20 (T) = PERTIMBANGAN DIVISI (HAIRCUT)
    
    Matching dilakukan berdasarkan KODE EFEK di col 3 (C),
    bukan urutan baris, agar aman jika urutan berbeda.
    """
    output = BytesIO()
    wb = load_workbook(file_template)
    ws_conc = wb["CONC"]
    ws_hc = wb["HC"]

    # Buat lookup dict dari df_hasil: kode_efek -> row dict
    hasil_map = {
        str(row.get('KODE EFEK', '')).strip(): row
        for row in df_hasil.to_dict('records')
    }

    # --- INJECT SHEET CONC ---
    for excel_row in ws_conc.iter_rows(min_row=5):
        kode = excel_row[2].value  # col C = index 2
        if kode is None:
            continue
        kode = str(kode).strip()
        if kode not in hasil_map:
            continue
        row = hasil_map[kode]
        ws_conc.cell(row=excel_row[0].row, column=16, value=row.get('CONCENTRATION LIMIT KARENA SAHAM MARJIN BARU'))
        ws_conc.cell(row=excel_row[0].row, column=17, value=row.get(COL_LISTED))
        ws_conc.cell(row=excel_row[0].row, column=18, value=row.get(COL_FF))
        ws_conc.cell(row=excel_row[0].row, column=19, value=row.get(COL_RMCC))
        ws_conc.cell(row=excel_row[0].row, column=22, value=row.get('PERTIMBANGAN DIVISI (CONC LIMIT)'))

    # --- INJECT SHEET HC ---
    for excel_row in ws_hc.iter_rows(min_row=5):
        kode = excel_row[2].value  # col C = index 2
        if kode is None:
            continue
        kode = str(kode).strip()
        if kode not in hasil_map:
            continue
        row = hasil_map[kode]
        ws_hc.cell(row=excel_row[0].row, column=18, value=row.get('HAIRCUT PEI USULAN DIVISI'))
        ws_hc.cell(row=excel_row[0].row, column=20, value=row.get('PERTIMBANGAN DIVISI (HAIRCUT)'))

    wb.save(output)
    output.seek(0)
    return output

# ============================
# ANTARMUKA STREAMLIT
# ============================

def main():
    st.title("🛡️ Concentration Limit (CL) & Haircut Calculation")

    tab1, tab2 = st.tabs(["📊 Hitung CL & Haircut", "💉 Inject ke Template"])

    # ============================================================
    # TAB 1: HITUNG
    # ============================================================
    with tab1:
        st.markdown("Unggah file sumber raw data untuk menjalankan perhitungan CL & Haircut.")

        current_month_name = datetime.now().strftime('%B').lower()
        example_filename = f'HCCL_{current_month_name}.xlsx'

        uploaded_file_cl = st.file_uploader(
            f"📂 Unggah File Sumber (misal: `{example_filename}`)",
            type=['xlsx'], key='cl_source'
        )

        if uploaded_file_cl is not None:
            if st.button("🚀 Jalankan Perhitungan CL", type="primary"):
                try:
                    df_cl_source = pd.read_excel(uploaded_file_cl, engine='openpyxl')

                    with st.spinner('Menghitung Concentration Limit...'):
                        df_cl_hasil = calculate_concentration_limit(df_cl_source)

                    st.success("✅ Perhitungan selesai!")
                    st.subheader("Hasil (Tabel)")
                    st.dataframe(df_cl_hasil)

                    output_buffer_cl = BytesIO()
                    df_cl_hasil.to_excel(output_buffer_cl, index=False)
                    output_buffer_cl.seek(0)

                    st.download_button(
                        label="⬇️ Unduh Hasil (Excel)",
                        data=output_buffer_cl,
                        file_name=f'Hasil Raw Data HCCL_{current_month_name}.xlsx',
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )

                except Exception as e:
                    st.error(f"❌ Gagal. Pastikan format file benar. Error: {e}")

    # ============================================================
    # TAB 2: INJECT
    # ============================================================
    with tab2:
        st.markdown("Unggah hasil perhitungan dari Tab 1 dan template target untuk inject.")

        col1, col2 = st.columns(2)
        with col1:
            uploaded_hasil = st.file_uploader(
                "📂 Unggah Hasil Perhitungan (dari Tab 1)",
                type=['xlsx'], key='cl_hasil'
            )
        with col2:
            uploaded_template = st.file_uploader(
                "📋 Unggah Template Target (XLSX)",
                type=['xlsx'], key='cl_template'
            )

        if uploaded_hasil is not None and uploaded_template is not None:
            if st.button("💉 Inject ke Template", type="primary"):
                try:
                    df_hasil = pd.read_excel(uploaded_hasil, engine='openpyxl')

                    with st.spinner('Menyuntikkan data ke template...'):
                        uploaded_template.seek(0)
                        final_xlsx = update_excel_template(uploaded_template, df_hasil)

                    st.success("✅ Inject selesai!")

                    current_month_name = datetime.now().strftime('%B').lower()
                    st.download_button(
                        label="⬇️ Unduh Hasil (Injected ke Template)",
                        data=final_xlsx,
                        file_name=f'Hasil_Template_{current_month_name}.xlsx',
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )

                except Exception as e:
                    st.error(f"❌ Gagal inject. Error: {e}")
        else:
            st.info("💡 Upload kedua file di atas untuk mengaktifkan tombol inject.")

if __name__ == '__main__':
    main()
