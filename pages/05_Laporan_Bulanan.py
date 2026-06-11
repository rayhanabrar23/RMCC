import streamlit as st
import pandas as pd
import os
import glob
import re
import warnings
import io
import zipfile
import tempfile
from openpyxl import load_workbook
import openpyxl

warnings.simplefilter(action="ignore", category=FutureWarning)

# ─────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────
st.set_page_config(
    page_title="RMCC | Laporan Bulanan",
    page_icon="📊",
    layout="wide",
)

# ─────────────────────────────────────────
# CUSTOM CSS
# ─────────────────────────────────────────
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

    .stApp { background-color: #0f1117; color: #e8eaf0; }

    /* Sidebar */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #1a1d2e 0%, #141624 100%);
        border-right: 1px solid #2a2d3e;
    }

    /* Tab bar */
    .stTabs [data-baseweb="tab-list"] {
        background: #1a1d2e;
        border-radius: 10px;
        padding: 4px;
        gap: 4px;
        border: 1px solid #2a2d3e;
    }
    .stTabs [data-baseweb="tab"] {
        color: #8a8fa8;
        font-weight: 500;
        font-size: 0.85rem;
        border-radius: 7px;
        padding: 8px 18px;
    }
    .stTabs [aria-selected="true"] {
        background: linear-gradient(135deg, #3b5bdb, #5c7cfa) !important;
        color: white !important;
    }

    /* Cards */
    .card {
        background: #1a1d2e;
        border: 1px solid #2a2d3e;
        border-radius: 12px;
        padding: 20px 24px;
        margin-bottom: 16px;
    }
    .card-title {
        font-size: 0.75rem;
        font-weight: 600;
        letter-spacing: 0.08em;
        color: #5c7cfa;
        text-transform: uppercase;
        margin-bottom: 6px;
    }

    /* Log box */
    .log-box {
        background: #0d1117;
        border: 1px solid #2a2d3e;
        border-radius: 8px;
        padding: 16px;
        font-family: 'Courier New', monospace;
        font-size: 0.78rem;
        line-height: 1.6;
        max-height: 300px;
        overflow-y: auto;
        color: #c9d1d9;
    }
    .log-ok   { color: #3fb950; }
    .log-err  { color: #f85149; }
    .log-warn { color: #d29922; }
    .log-info { color: #58a6ff; }

    /* Status badge */
    .badge {
        display: inline-block;
        padding: 2px 10px;
        border-radius: 20px;
        font-size: 0.72rem;
        font-weight: 600;
    }
    .badge-blue   { background: #1c2d5e; color: #5c7cfa; }
    .badge-green  { background: #0d3321; color: #3fb950; }
    .badge-red    { background: #3d0c0c; color: #f85149; }

    /* Header */
    .main-header {
        background: linear-gradient(135deg, #1a1d2e 0%, #1e2340 100%);
        border: 1px solid #2a2d3e;
        border-radius: 14px;
        padding: 24px 28px;
        margin-bottom: 24px;
    }
    .main-header h1 {
        font-size: 1.6rem;
        font-weight: 700;
        color: #e8eaf0;
        margin: 0;
    }
    .main-header p {
        color: #6b7080;
        margin: 6px 0 0 0;
        font-size: 0.88rem;
    }

    /* Buttons */
    .stButton > button {
        background: linear-gradient(135deg, #3b5bdb, #5c7cfa);
        color: white;
        border: none;
        border-radius: 8px;
        font-weight: 600;
        font-size: 0.85rem;
        padding: 10px 24px;
        transition: opacity 0.2s;
    }
    .stButton > button:hover { opacity: 0.88; }

    /* Upload area */
    [data-testid="stFileUploader"] {
        background: #1a1d2e;
        border: 1px dashed #3b5bdb;
        border-radius: 10px;
        padding: 8px;
    }

    /* Metric */
    [data-testid="stMetric"] {
        background: #1a1d2e;
        border: 1px solid #2a2d3e;
        border-radius: 10px;
        padding: 14px 18px;
    }
    [data-testid="stMetricLabel"] { color: #8a8fa8 !important; font-size: 0.78rem !important; }
    [data-testid="stMetricValue"] { color: #e8eaf0 !important; font-size: 1.2rem !important; }

    /* Divider */
    hr { border-color: #2a2d3e; }

    /* Select/Input */
    .stSelectbox > div > div, .stTextInput > div > div {
        background: #1a1d2e !important;
        border-color: #2a2d3e !important;
        color: #e8eaf0 !important;
    }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────
# SIDEBAR – KONFIGURASI GLOBAL
# ─────────────────────────────────────────
with st.sidebar:
    st.markdown("### ⚙️ Konfigurasi")
    st.markdown("---")

    BULAN_LIST = ["Januari","Februari","Maret","April","Mei","Juni",
                  "Juli","Agustus","September","Oktober","November","Desember"]
    bulan = st.selectbox("Bulan", BULAN_LIST, index=4)
    tahun = st.text_input("Tahun", value="2026")

    st.markdown("---")
    st.markdown('<span class="badge badge-blue">v1.0</span> &nbsp; RMCC Automation', unsafe_allow_html=True)
    st.markdown('<p style="color:#4a5060;font-size:0.75rem;margin-top:8px;">Semua proses berjalan di memori.<br>File output bisa diunduh langsung.</p>', unsafe_allow_html=True)

# ─────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────
st.markdown(f"""
<div class="main-header">
  <h1>📊 RMCC — Otomasi Laporan Bulanan</h1>
  <p>Periode aktif: <strong style="color:#5c7cfa">{bulan} {tahun}</strong> &nbsp;·&nbsp; Pilih tab untuk memproses laporan</p>
</div>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────
# HELPER
# ─────────────────────────────────────────
def log_line(msg: str, kind: str = "info") -> str:
    icons = {"ok": "✅", "err": "❌", "warn": "⚠️", "info": "ℹ️"}
    css   = {"ok": "log-ok", "err": "log-err", "warn": "log-warn", "info": "log-info"}
    icon  = icons.get(kind, "·")
    cls   = css.get(kind, "log-info")
    return f'<span class="{cls}">{icon} {msg}</span><br>'


def render_log(lines: list):
    html = '<div class="log-box">' + "".join(lines) + "</div>"
    st.markdown(html, unsafe_allow_html=True)


def to_excel_bytes(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def smart_read_excel(uploaded_file):
    try:
        return pd.read_excel(uploaded_file, header=None)
    except Exception:
        try:
            dfs = pd.read_html(uploaded_file)
            return dfs[0] if dfs else None
        except Exception:
            return None


# ─────────────────────────────────────────
# TABS
# ─────────────────────────────────────────
tab1, tab2, tab3, tab4 = st.tabs([
    "💰 Disburse & Repayment",
    "👥 Participant & Client",
    "📦 Collateral Value",
    "📂 Margin Trades CSV",
])

# ══════════════════════════════════════════
# TAB 1 — DISBURSE & REPAYMENT
# ══════════════════════════════════════════
with tab1:
    st.markdown('<div class="card"><div class="card-title">Upload File</div>', unsafe_allow_html=True)
    st.markdown("Upload **file master** dan semua **file Disburse/Repay** per broker (EP, HD, HP, XC).")

    col_a, col_b = st.columns([1, 1])
    with col_a:
        master_file = st.file_uploader(
            "File Master (Disburse dan Repayment Bulan Tahun.xlsx)",
            type=["xlsx"], key="t1_master"
        )
    with col_b:
        broker_files = st.file_uploader(
            "File Broker (Disburse & Repay per EP/HD/HP/XC) — bisa multi-file",
            type=["xlsx", "xls"], accept_multiple_files=True, key="t1_broker"
        )
    st.markdown("</div>", unsafe_allow_html=True)

    if st.button("▶ Proses Disburse & Repayment", key="run_t1"):
        if not master_file or not broker_files:
            st.warning("Upload file master dan minimal 1 file broker dulu ya.")
        else:
            logs = []
            MONTH_COL = {
                'JANUARI':'F','FEBRUARI':'G','MARET':'H','APRIL':'J',
                'MEI':'K','JUNI':'L','JULI':'N','AGUSTUS':'O',
                'SEPTEMBER':'P','OKTOBER':'R','NOVEMBER':'S','DESEMBER':'T'
            }
            target_col = MONTH_COL.get(bulan.upper(), 'F')
            row_disburse   = {'EP':5,'HD':6,'HP':7,'XC':8}
            row_repayment  = {'EP':16,'HD':17,'HP':18,'XC':19}

            try:
                wb = load_workbook(master_file)
                ws = wb['Disburse & Repay Jan-Des']
                logs.append(log_line(f"File master dibuka, target kolom: {target_col}", "ok"))
            except Exception as e:
                logs.append(log_line(f"Gagal buka file master: {e}", "err"))
                render_log(logs)
                st.stop()

            processed = 0
            for uf in broker_files:
                fname = uf.name.upper()
                if "~$" in fname:
                    continue

                words = re.findall(r'[A-Z0-9]+', fname)
                code = next((c for c in ['EP','HD','HP','XC'] if c in words), None)
                is_disburse = "DISB" in fname
                is_repay    = "REPAY" in fname or "RPY" in fname

                if not code or (not is_disburse and not is_repay):
                    logs.append(log_line(f"Skip: {uf.name} (nama tidak sesuai standar)", "warn"))
                    continue

                df_raw = smart_read_excel(uf)
                if df_raw is None:
                    logs.append(log_line(f"Gagal membaca: {uf.name}", "err"))
                    continue

                try:
                    val = 0
                    df_str = df_raw.astype(str)

                    if is_disburse:
                        mask = df_str.apply(
                            lambda x: x.str.contains('Grand Total', case=False, na=False)
                        ).any(axis=1)
                        df_total = df_raw[mask]
                        if not df_total.empty:
                            numeric = pd.to_numeric(df_total.iloc[-1], errors='coerce').dropna()
                            val = numeric.iloc[-1] if not numeric.empty else 0
                        ws[f"{target_col}{row_disburse[code]}"] = val
                        logs.append(log_line(f"[DISBURSE] {code}: {val:,.2f}  ← {uf.name}", "ok"))
                        processed += 1

                    elif is_repay:
                        numeric_df = df_raw.apply(pd.to_numeric, errors='coerce')
                        df_clean = numeric_df.dropna(axis=1, how='all')
                        if not df_clean.empty:
                            last_col = df_clean.columns[-1]
                            series = df_clean[last_col].dropna()
                            val = series.iloc[-1] if not series.empty else 0
                        ws[f"{target_col}{row_repayment[code]}"] = val
                        logs.append(log_line(f"[REPAYMENT] {code}: {val:,.2f}  ← {uf.name}", "ok"))
                        processed += 1

                except Exception as e:
                    logs.append(log_line(f"Error {uf.name}: {e}", "err"))

            # Update Summary
            try:
                summary_row = BULAN_LIST.index(bulan) + 2
                ws_sum = wb['Summary Tahunan']
                ws_sum[f'B{summary_row}'] = f"=SUM('Disburse & Repay Jan-Des'!{target_col}5:{target_col}8)"
                logs.append(log_line(f"Summary Tahunan diperbarui (baris {summary_row})", "ok"))
            except Exception as e:
                logs.append(log_line(f"Gagal update Summary: {e}", "warn"))

            render_log(logs)

            col1, col2 = st.columns([1, 3])
            with col1:
                st.metric("File Diproses", processed)
            with col2:
                st.download_button(
                    label=f"⬇ Unduh File Master ({bulan} {tahun})",
                    data=to_excel_bytes(wb),
                    file_name=f"Disburse dan Repayment {bulan} {tahun}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )


# ══════════════════════════════════════════
# TAB 2 — PARTICIPANT & CLIENT POSITION
# ══════════════════════════════════════════
with tab2:
    st.markdown('<div class="card"><div class="card-title">Upload File</div>', unsafe_allow_html=True)
    st.markdown("Upload file **Participant And Client Detail Position** untuk bulan ini.")

    t2_file = st.file_uploader(
        f"Participant And Client Detail Position {bulan} {tahun}.xlsx",
        type=["xlsx"], key="t2_input"
    )
    repo_value = st.number_input(
        "Outstanding REPO (Rp)", value=181_289_543_803,
        step=1_000_000, format="%d", key="t2_repo"
    )
    st.markdown("</div>", unsafe_allow_html=True)

    if st.button("▶ Proses Participant & Client", key="run_t2"):
        if not t2_file:
            st.warning("Upload file dulu ya.")
        else:
            logs = []
            broker_codes = ['EP','HP','HD','XC']
            rename_map   = {'EP':'MNC','HP':'Henan','HD':'KGI','XC':'AJAIB'}

            try:
                df_raw = pd.read_excel(t2_file, skiprows=1)
                df_raw.columns = [str(c).strip() for c in df_raw.columns]
                cols = df_raw.columns.tolist()
                mapping = {
                    'date':cols[0],'broker':cols[2],'name':cols[3],
                    'mkbd':cols[4],'os':cols[5],'limit':cols[6],
                    'avail':cols[7],'perc':cols[8],'int':cols[9],'short':cols[10]
                }
                df_raw['Value Date'] = pd.to_datetime(df_raw[mapping['date']], errors='coerce')
                df_raw = df_raw.dropna(subset=['Value Date'])
                logs.append(log_line(f"Data dibaca: {len(df_raw):,} baris", "ok"))
            except Exception as e:
                logs.append(log_line(f"Gagal membaca file: {e}", "err"))
                render_log(logs)
                st.stop()

            output_buf = io.BytesIO()
            all_summaries, existing_brokers, last_positions = [], [], []

            # Simpan sheet data asli
            with pd.ExcelWriter(output_buf, engine='openpyxl') as writer:
                df_raw.to_excel(writer, sheet_name=f'Data {bulan}', index=False)

                for code in broker_codes:
                    df_b = df_raw[df_raw[mapping['broker']].astype(str).str.contains(code, na=False)].copy()
                    if df_b.empty:
                        logs.append(log_line(f"Broker {code}: tidak ada data", "warn"))
                        continue

                    df_daily = df_b.groupby(['Value Date', mapping['broker'], mapping['name']]).agg({
                        mapping['mkbd']:'first', mapping['os']:'first', mapping['limit']:'first',
                        mapping['avail']:'first', mapping['perc']:'first',
                        mapping['int']:'first', mapping['short']:'first'
                    }).reset_index().sort_values('Value Date')

                    last_positions.append(df_daily.tail(1))

                    df_sum = df_daily[['Value Date', mapping['os']]].copy()
                    df_sum.columns = ['Tanggal', code]
                    all_summaries.append(df_sum)
                    existing_brokers.append(code)

                    df_disp = df_daily.copy()
                    df_disp['Value Date'] = df_disp['Value Date'].dt.strftime('%d %B %Y')
                    df_disp.to_excel(writer, sheet_name=f'Report Part {code}', index=False)
                    logs.append(log_line(f"Sheet Report Part {code} dibuat ({len(df_daily)} hari)", "ok"))

                if all_summaries:
                    df_final = all_summaries[0]
                    for i in range(1, len(all_summaries)):
                        df_final = pd.merge(df_final, all_summaries[i], on='Tanggal', how='outer')
                    df_final = df_final.sort_values('Tanggal').fillna(0)
                    cols_p = [c for c in existing_brokers if c in df_final.columns]
                    df_final['Total Marjin'] = df_final[cols_p].sum(axis=1)
                    df_final['Outstanding REPO'] = repo_value

                    # Rekap
                    df_rekap = df_final.copy()
                    df_rekap['Tanggal'] = df_rekap['Tanggal'].dt.strftime('%d %B %Y')
                    df_rekap_display = df_rekap.rename(columns=rename_map)
                    df_rekap_display.to_excel(writer, sheet_name='Rekap Outstanding', index=False)
                    logs.append(log_line("Sheet Rekap Outstanding dibuat", "ok"))

                    # T1 Pertumbuhan Bulanan
                    t1_list = []
                    for col in cols_p + ['Total Marjin']:
                        v0 = df_final[col].iloc[0]
                        v1 = df_final[col].iloc[-1]
                        g  = (v1-v0)/v0 if v0 != 0 else 0
                        t1_list.append([df_final[col].max(), df_final[col].min(), g])
                    df_t1 = pd.DataFrame(
                        t1_list,
                        index=[rename_map.get(c,c) for c in cols_p]+['Total'],
                        columns=['Max','Min','Pertumbuhan']
                    ).T
                    df_t1.to_excel(writer, sheet_name='T1-T2 Pertumbuhan Bulanan')
                    logs.append(log_line("Sheet T1-T2 Pertumbuhan Bulanan dibuat", "ok"))

                    # T6 Pertumbuhan Harian
                    df_t6 = df_final[['Tanggal']+cols_p+['Total Marjin']].copy()
                    df_t6['Pertumbuhan'] = (
                        df_t6['Total Marjin'].shift(-1) - df_t6['Total Marjin']
                    ) / df_t6['Total Marjin']
                    df_t6['Tanggal'] = df_t6['Tanggal'].dt.strftime('%d %B %Y')
                    df_t6.to_excel(writer, sheet_name='T6 Pertumbuhan Harian', index=False)
                    logs.append(log_line("Sheet T6 Pertumbuhan Harian dibuat", "ok"))

                    # Data Per Akhir Bulan
                    df_last = pd.concat(last_positions, ignore_index=True)
                    total_os = df_last[mapping['os']].sum()
                    df_last['% Outstanding'] = df_last[mapping['os']] / total_os if total_os else 0
                    cols_order = [
                        mapping['date'], mapping['broker'], mapping['name'],
                        mapping['mkbd'], mapping['os'], '% Outstanding',
                        mapping['limit'], mapping['avail'], mapping['perc'],
                        mapping['int'], mapping['short']
                    ]
                    df_last = df_last[cols_order]
                    df_last[mapping['date']] = df_last[mapping['date']].dt.strftime('%d %B %Y')
                    df_last.to_excel(writer, sheet_name='Data Per Akhir Bulan', index=False)

                    ws_last = writer.sheets['Data Per Akhir Bulan']
                    last_row = len(df_last) + 2
                    ws_last.cell(row=last_row, column=4).value = "TOTAL"
                    ws_last.cell(row=last_row, column=5).value = total_os
                    ws_last.cell(row=last_row, column=6).value = 1.0
                    ws_last.cell(row=last_row, column=6).number_format = '0.00%'
                    logs.append(log_line("Sheet Data Per Akhir Bulan dibuat", "ok"))

            render_log(logs)

            col1, col2 = st.columns([1, 3])
            with col1:
                st.metric("Sheet Dibuat", len(existing_brokers) + 4)
            with col2:
                st.download_button(
                    label=f"⬇ Unduh Laporan ({bulan} {tahun})",
                    data=output_buf.getvalue(),
                    file_name=f"Format_Participant_Client_{bulan}_{tahun}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )


# ══════════════════════════════════════════
# TAB 3 — COLLATERAL VALUE
# ══════════════════════════════════════════
with tab3:
    st.markdown('<div class="card"><div class="card-title">Upload File</div>', unsafe_allow_html=True)
    st.markdown("Upload **Stock Position Detail** dan **Collateral Value** (file rekap tahunan).")

    col_a, col_b = st.columns(2)
    with col_a:
        t3_stock = st.file_uploader(
            f"Stock Position Detail {bulan} {tahun}.xlsx",
            type=["xlsx"], key="t3_stock"
        )
    with col_b:
        t3_collateral = st.file_uploader(
            f"Collateral Value {tahun}.xlsx (file rekap tahunan)",
            type=["xlsx"], key="t3_coll"
        )
    st.markdown("</div>", unsafe_allow_html=True)

    if st.button("▶ Proses Collateral Value", key="run_t3"):
        if not t3_stock or not t3_collateral:
            st.warning("Upload kedua file dulu ya.")
        else:
            logs = []
            try:
                df = pd.read_excel(t3_stock)
                df.columns = [str(c).strip() for c in df.columns]

                col_index = next(
                    (c for c in ['Index','Indeks'] if c in df.columns),
                    df.columns[7]
                )
                col_value = next(
                    (c for c in df.columns if 'Collateral' in str(c)), None
                )

                if not col_value:
                    logs.append(log_line("Kolom Collateral Value tidak ditemukan!", "err"))
                    render_log(logs)
                    st.stop()

                df = df[~df.iloc[:,0].astype(str).str.contains('Total', case=False, na=False)]

                def label_group(val):
                    val = str(val).strip()
                    if val == "IHSG,IDX80,LQ45": return "LQ45"
                    elif val == "IHSG,IDX80":    return "IDX80"
                    else:                         return "MARJIN"

                df['Group']  = df[col_index].apply(label_group)
                df[col_value] = pd.to_numeric(df[col_value], errors='coerce').fillna(0)

                summary     = df.groupby(col_index)[col_value].sum()
                val_lq45    = summary.get("IHSG,IDX80,LQ45", 0)
                val_idx80   = summary.get("IHSG,IDX80", 0)
                val_marjin  = summary.get("IHSG", 0)

                logs.append(log_line(f"LQ45   : {val_lq45:,.0f}", "ok"))
                logs.append(log_line(f"IDX80  : {val_idx80:,.0f}", "ok"))
                logs.append(log_line(f"Marjin : {val_marjin:,.0f}", "ok"))

            except Exception as e:
                logs.append(log_line(f"Gagal proses Stock Detail: {e}", "err"))
                render_log(logs)
                st.stop()

            # Tulis ke file Collateral Value (rekap tahunan)
            try:
                wb = load_workbook(t3_collateral)

                # Simpan sheet detail bulan ini
                buf_detail = io.BytesIO()
                with pd.ExcelWriter(buf_detail, engine='openpyxl') as wr:
                    df.to_excel(wr, sheet_name=bulan, index=False)
                    summary.reset_index().to_excel(wr, sheet_name=f'Pivot_{bulan}', index=False)

                # Masukkan ke sheet Recap
                ws_recap = wb['Collateral Value Recap']
                target_row = 3
                while ws_recap[f'B{target_row}'].value is not None:
                    target_row += 1

                ws_recap[f'A{target_row}'] = f"{bulan} {tahun}"
                ws_recap[f'B{target_row}'] = val_lq45
                ws_recap[f'C{target_row}'] = val_idx80
                ws_recap[f'D{target_row}'] = val_marjin
                logs.append(log_line(f"Data ditulis ke Recap baris {target_row}", "ok"))

                # Tambah sheet detail ke wb yang sama
                wb_detail = load_workbook(buf_detail)
                for sname in wb_detail.sheetnames:
                    ws_src = wb_detail[sname]
                    if sname in wb.sheetnames:
                        del wb[sname]
                    ws_new = wb.create_sheet(sname)
                    for row in ws_src.iter_rows():
                        for cell in row:
                            ws_new[cell.coordinate].value = cell.value

                logs.append(log_line(f"Sheet '{bulan}' dan 'Pivot_{bulan}' ditambahkan", "ok"))

            except Exception as e:
                logs.append(log_line(f"Gagal update file Collateral: {e}", "err"))
                render_log(logs)
                st.stop()

            render_log(logs)

            col1, col2, col3, col4 = st.columns(4)
            col1.metric("LQ45",   f"{val_lq45:,.0f}")
            col2.metric("IDX80",  f"{val_idx80:,.0f}")
            col3.metric("Marjin", f"{val_marjin:,.0f}")
            col4.metric("Total",  f"{val_lq45+val_idx80+val_marjin:,.0f}")

            st.download_button(
                label=f"⬇ Unduh Collateral Value {tahun} (Updated)",
                data=to_excel_bytes(wb),
                file_name=f"Collateral Value {tahun}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )


# ══════════════════════════════════════════
# TAB 4 — MARGIN TRADES CSV
# ══════════════════════════════════════════
with tab4:
    st.markdown('<div class="card"><div class="card-title">Upload File</div>', unsafe_allow_html=True)
    st.markdown("Upload semua file **MarginTrades_\\*.csv** harian sekaligus. Otomatis digabung & di-split jika >1 juta baris.")

    t4_files = st.file_uploader(
        "MarginTrades_*.csv (multi-file)",
        type=["csv"], accept_multiple_files=True, key="t4_csv"
    )
    limit_baris = st.number_input(
        "Batas split (baris)", value=1_000_000,
        step=100_000, format="%d", key="t4_limit"
    )
    st.markdown("</div>", unsafe_allow_html=True)

    if st.button("▶ Gabung & Proses CSV", key="run_t4"):
        if not t4_files:
            st.warning("Upload file CSV dulu ya.")
        else:
            logs = []
            all_data = []

            for f in sorted(t4_files, key=lambda x: x.name):
                try:
                    df = pd.read_csv(
                        f, sep=None, engine='python',
                        header=None, skiprows=1
                    )
                    all_data.append(df.iloc[:, 0:13])
                    logs.append(log_line(f"{f.name}: {len(df):,} baris", "ok"))
                except Exception as e:
                    logs.append(log_line(f"Gagal baca {f.name}: {e}", "err"))

            if not all_data:
                logs.append(log_line("Tidak ada data yang berhasil dibaca.", "err"))
                render_log(logs)
                st.stop()

            final_df = pd.concat(all_data, ignore_index=True)
            final_df.columns = list('ABCDEFGHIJKLM')
            total_baris = len(final_df)
            logs.append(log_line(f"Total gabungan: {total_baris:,} baris dari {len(t4_files)} file", "info"))

            base_name = f"Updated_{tahun}_Q1_Margin_{bulan}"

            render_log(logs)
            st.metric("Total Baris", f"{total_baris:,}")

            if total_baris > limit_baris:
                st.info(f"Data melebihi {limit_baris:,} baris → dipecah jadi 2 file.")

                buf1 = io.BytesIO()
                buf2 = io.BytesIO()
                final_df.iloc[:limit_baris].to_csv(buf1, index=False, sep=';')
                final_df.iloc[limit_baris:].to_csv(buf2, index=False, sep=';')

                col1, col2 = st.columns(2)
                with col1:
                    st.download_button(
                        label=f"⬇ Unduh Part 1 (baris 1–{limit_baris:,})",
                        data=buf1.getvalue(),
                        file_name=f"{base_name}_Part1.csv",
                        mime="text/csv",
                        use_container_width=True,
                    )
                with col2:
                    st.download_button(
                        label=f"⬇ Unduh Part 2 (baris {limit_baris+1:,}–{total_baris:,})",
                        data=buf2.getvalue(),
                        file_name=f"{base_name}_Part2.csv",
                        mime="text/csv",
                        use_container_width=True,
                    )
            else:
                buf = io.BytesIO()
                final_df.to_csv(buf, index=False, sep=';')
                st.download_button(
                    label=f"⬇ Unduh {base_name}.csv",
                    data=buf.getvalue(),
                    file_name=f"{base_name}.csv",
                    mime="text/csv",
                    use_container_width=True,
                )
