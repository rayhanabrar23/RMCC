import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

st.set_page_config(page_title="Collateral Value Processor", layout="wide")

st.title("📊 Collateral Value & Recap Automator")
st.markdown("Proses data *Stock Position Detail* ke dalam *Collateral Value Master* secara otomatis.")

# --- SIDEBAR KONFIGURASI ---
with st.sidebar:
    st.header("Konfigurasi Periode")
    bulan_proses = st.selectbox("Pilih Bulan", 
                                ["Januari", "Februari", "Maret", "April", "Mei", "Juni", 
                                 "Juli", "Agustus", "September", "Oktober", "November", "Desember"])
    tahun_proses = st.text_input("Tahun", "2026")

# --- UPLOAD FILE ---
col1, col2 = st.columns(2)

with col1:
    file_satu = st.file_uploader("1. Unggah Stock Position Detail (Excel)", type=['xlsx'])

with col2:
    file_dua_master = st.file_uploader("2. Unggah Master Collateral Value (Recap)", type=['xlsx'])

def process_data(stock_file, master_file, bulan, tahun):
    try:
        # 1. Membaca file Stock Position
        df = pd.read_excel(stock_file)
        df.columns = [str(c).strip() for c in df.columns]

        # Identifikasi Kolom
        col_index = next((c for c in ['Index', 'Indeks'] if c in df.columns), df.columns[7])
        col_value = next((c for c in df.columns if 'Collateral' in str(c)), None)

        if not col_value:
            return None, "Kolom 'Collateral Value' tidak ditemukan dalam file sumber!"

        # 2. Pembersihan Data
        df = df[~df.iloc[:, 0].astype(str).str.contains('Total', case=False, na=False)]
        df[col_value] = pd.to_numeric(df[col_value], errors='coerce').fillna(0)

        # 3. Hitung Pivot
        summary = df.groupby(col_index)[col_value].sum()
        val_lq45 = summary.get("IHSG,IDX80,LQ45", 0)
        val_idx80 = summary.get("IHSG,IDX80", 0)
        val_marjin = summary.get("IHSG", 0)

        # 4. Proses Writing ke Excel Master (di Memori)
        # Kita baca master_file ke memory
        output = BytesIO()
        output.write(master_file.getvalue())
        
        # Gunakan openpyxl untuk recap (baris kosong)
        wb = load_workbook(output)
        if 'Collateral Value Recap' not in wb.sheetnames:
            return None, "Sheet 'Collateral Value Recap' tidak ditemukan di file Master!"
            
        ws_recap = wb['Collateral Value Recap']

        # Mencari baris pertama yang kosong di kolom B (mulai dari baris 3)
        target_row = 3
        while ws_recap[f'B{target_row}'].value is not None:
            target_row += 1

        # Isi data
        ws_recap[f'B{target_row}'] = val_lq45
        ws_recap[f'C{target_row}'] = val_idx80
        ws_recap[f'D{target_row}'] = val_marjin
        
        if ws_recap[f'A{target_row}'].value is None:
            ws_recap[f'A{target_row}'] = f"{bulan} {tahun}"

        # Simpan ke memori untuk diunduh
        final_output = BytesIO()
        
        # Gunakan Pandas Writer untuk menambah sheet detail bulan ini
        with pd.ExcelWriter(final_output, engine='openpyxl') as writer:
            writer.book = wb
            # Tambahkan sheet detail
            df.to_excel(writer, sheet_name=bulan, index=False)
            # Tambahkan sheet pivot
            summary.reset_index().to_excel(writer, sheet_name=f'Pivot_{bulan}', index=False)
        
        final_output.seek(0)
        
        log_data = {
            "Bulan": bulan,
            "LQ45": val_lq45,
            "IDX80": val_idx80,
            "MARJIN": val_marjin,
            "Baris Recap": target_row
        }
        
        return final_output, log_data

    except Exception as e:
        return None, str(e)

# --- TOMBOL EKSEKUSI ---
if file_satu and file_dua_master:
    if st.button("🚀 Proses Data"):
        with st.spinner("Sedang memproses..."):
            result_file, message = process_data(file_satu, file_dua_master, bulan_proses, tahun_proses)
            
            if result_file:
                st.success(f"Berhasil memproses data {bulan_proses}!")
                
                # Tampilkan Ringkasan
                st.write("### 📊 Ringkasan Hasil:")
                st.table(pd.DataFrame([message]))
                
                # Tombol Download
                st.download_button(
                    label="📥 Download File Collateral Update",
                    data=result_file,
                    file_name=f"Collateral Value {tahun_proses} - Update {bulan_proses}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.error(f"Terjadi kesalahan: {message}")

else:
    st.info("Silakan unggah kedua file Excel di atas untuk memulai pemrosesan.")
