import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, NamedStyle
from io import BytesIO
from datetime import datetime
import os
import sys

from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch

# Cek apakah sudah login dari halaman utama
if "login_status" not in st.session_state or not st.session_state["login_status"]:
    st.error("🚨 Akses Ditolak! Silakan login di halaman utama terlebih dahulu.")
    st.stop()

# ─────────────────────────────────────────
# CUSTOM CSS (sama dengan laporan bulanan)
# ─────────────────────────────────────────
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

    .stApp { background-color: #0f1117; color: #e8eaf0; }

    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #1a1d2e 0%, #141624 100%);
        border-right: 1px solid #2a2d3e;
    }

    .stTabs [data-baseweb="tab-list"] {
        background: #1a1d2e;
        border-radius: 10px;
        padding: 4px;
        gap: 4px;
        border: 1px solid #2a2d3e;
    }
    .stTabs [data-baseweb="tab"] {
        color: #8a8fa8;
        font-weight: 500;
        font-size: 0.85rem;
        border-radius: 7px;
        padding: 8px 18px;
    }
    .stTabs [aria-selected="true"] {
        background: linear-gradient(135deg, #3b5bdb, #5c7cfa) !important;
        color: white !important;
    }

    .card {
        background: #1a1d2e;
        border: 1px solid #2a2d3e;
        border-radius: 12px;
        padding: 20px 24px;
        margin-bottom: 16px;
    }
    .card-title {
        font-size: 0.75rem;
        font-weight: 600;
        letter-spacing: 0.08em;
        color: #5c7cfa;
        text-transform: uppercase;
        margin-bottom: 6px;
    }

    .log-box {
        background: #0d1117;
        border: 1px solid #2a2d3e;
        border-radius: 8px;
        padding: 16px;
        font-family: 'Courier New', monospace;
        font-size: 0.78rem;
        line-height: 1.6;
        max-height: 300px;
        overflow-y: auto;
        color: #c9d1d9;
    }
    .log-ok   { color: #3fb950; }
    .log-err  { color: #f85149; }
    .log-warn { color: #d29922; }
    .log-info { color: #58a6ff; }

    .badge {
        display: inline-block;
        padding: 2px 10px;
        border-radius: 20px;
        font-size: 0.72rem;
        font-weight: 600;
    }
    .badge-blue  { background: #1c2d5e; color: #5c7cfa; }
    .badge-green { background: #0d3321; color: #3fb950; }
    .badge-red   { background: #3d0c0c; color: #f85149; }

    .main-header {
        background: linear-gradient(135deg, #1a1d2e 0%, #1e2340 100%);
        border: 1px solid #2a2d3e;
        border-radius: 14px;
        padding: 24px 28px;
        margin-bottom: 24px;
    }
    .main-header h1 { font-size: 1.6rem; font-weight: 700; color: #e8eaf0; margin: 0; }
    .main-header p  { color: #6b7080; margin: 6px 0 0 0; font-size: 0.88rem; }

    .stButton > button {
        background: linear-gradient(135deg, #3b5bdb, #5c7cfa);
        color: white;
        border: none;
        border-radius: 8px;
        font-weight: 600;
        font-size: 0.85rem;
        padding: 10px 24px;
        transition: opacity 0.2s;
    }
    .stButton > button:hover { opacity: 0.88; }

    [data-testid="stFileUploader"] {
        background: #1a1d2e;
        border: 1px dashed #3b5bdb;
        border-radius: 10px;
        padding: 8px;
    }

    [data-testid="stMetric"] {
        background: #1a1d2e;
        border: 1px solid #2a2d3e;
        border-radius: 10px;
        padding: 14px 18px;
    }
    [data-testid="stMetricLabel"] { color: #8a8fa8 !important; font-size: 0.78rem !important; }
    [data-testid="stMetricValue"] { color: #e8eaf0 !important; font-size: 1.2rem !important; }

    hr { border-color: #2a2d3e; }

    .stSelectbox > div > div, .stTextInput > div > div {
        background: #1a1d2e !important;
        border-color: #2a2d3e !important;
        color: #e8eaf0 !important;
    }
</style>
""", unsafe_allow_html=True)

# ============================
# KONFIGURASI GLOBAL LL
# ============================
STOCK_CODE_BLACKLIST = ['BEBS', 'IPPE', 'WMPP', 'WMUU']
SHEET_INST_OLD = 'Instrument'
SHEET_INST_NEW = 'Hasil Pivot'
FIRST_LARGERST_COL = "First Largerst"
SECOND_LARGERST_COL = "Second Largerst"
SHEET_RESULT_NAME_SOURCE = 'Lendable Limit Result'
BORROW_AMOUNT_COL = 'Borrow Amount (shares)'

FINAL_COLUMNS_LL = [
    'Stock Code', 'Stock Name', 'Quantity On Hand',
    FIRST_LARGERST_COL, SECOND_LARGERST_COL, 'Total two Largerst',
    'Quantity Available', 'Thirty Percent On Hand', 'REPO',
    'Lendable Limit', 'Borrow Position', 'Available Lendable Limit'
]

# ============================
# FUNGSI STYLING KONDISIONAL
# ============================
def highlight_negative_ll(row):
    value = row['Available Lendable Limit']
    styles = [''] * len(row)
    if value < 0:
        styles = ['background-color: #FFCCCC'] * len(row)
    return styles

# ============================
# FUNGSI PEMROSESAN LL
# ============================
def process_lendable_limit(uploaded_files, template_file_data):
    st.info("🚀 Mulai Pemrosesan Lendable Limit...")
    try:
        df_sp = pd.read_excel(uploaded_files['Stock Position Detail.xlsx'], header=0, engine='openpyxl')
        df_instr_raw = pd.read_excel(uploaded_files['Instrument.xlsx'], sheet_name=SHEET_INST_OLD, header=1, engine='openpyxl')
        df_instr_old_raw = pd.read_excel(uploaded_files['Instrument.xlsx'], sheet_name=SHEET_INST_OLD, header=None, engine='openpyxl')
        df_borr_pos = pd.read_excel(uploaded_files['BorrPosition.xlsx'], header=0, engine='openpyxl')
    except Exception as e:
        st.error(f"❌ Gagal membaca salah satu file input LL. Error: {e}")
        return None, None, None

    output_xlsx_buffer = BytesIO()

    with st.spinner('Memproses data...'):
        try:
            df_sp.columns = df_sp.columns.str.strip()
            stock_col = df_sp.columns[1]
            qty_col = df_sp.columns[10]
            df_sp[qty_col] = pd.to_numeric(df_sp[qty_col], errors='coerce').fillna(0)

            top_values = (
                df_sp.groupby(stock_col)[qty_col]
                .apply(lambda x: sorted(x.dropna(), reverse=True)[:2])
                .reset_index()
            )
            top_values[FIRST_LARGERST_COL] = top_values[qty_col].apply(lambda x: x[0] if len(x) > 0 else 0)
            top_values[SECOND_LARGERST_COL] = top_values[qty_col].apply(lambda x: x[1] if len(x) > 1 else 0)
            df_sp = df_sp.merge(top_values.drop(columns=[qty_col]), how="left", on=stock_col).fillna({FIRST_LARGERST_COL: 0, SECOND_LARGERST_COL: 0})

            df_instr = df_instr_raw.copy()
            df_instr.columns = df_instr.columns.str.strip()
            col_row, col_loan, col_repo = "Local Code", "Used Loan Qty", "Used Reverse Repo Qty"
            df_instr[col_loan] = pd.to_numeric(df_instr[col_loan], errors='coerce').fillna(0)
            df_instr[col_repo] = pd.to_numeric(df_instr[col_repo], errors='coerce').fillna(0)

            df_pivot_full = df_instr.groupby(col_row)[[col_loan, col_repo]].sum().reset_index()

            df_result = df_pivot_full[['Local Code']].rename(columns={'Local Code': 'Stock Code'}).drop_duplicates()
            df_inst_lookup = df_instr_old_raw.iloc[1:].rename(columns={2: 'Stock Code', 9: 'Stock Name'})[['Stock Code', 'Stock Name']].drop_duplicates('Stock Code')
            df_result = df_result.merge(df_inst_lookup, on='Stock Code', how='left')

            qoh_calc = df_sp.groupby(df_sp.columns[1])[df_sp.columns[10]].sum().reset_index().rename(columns={df_sp.columns[1]: 'Stock Code', df_sp.columns[10]: 'Quantity On Hand'})
            df_result = df_result.merge(qoh_calc, on='Stock Code', how='left').fillna(0)

            borr_qty_calc = df_borr_pos.groupby('Stock Code')[BORROW_AMOUNT_COL].sum().reset_index().rename(columns={BORROW_AMOUNT_COL: 'Borrow Position'})
            df_result = df_result.merge(borr_qty_calc, on='Stock Code', how='left').fillna(0)

            largest_calc = df_sp.rename(columns={df_sp.columns[1]: 'Stock Code'}).groupby('Stock Code')[[FIRST_LARGERST_COL, SECOND_LARGERST_COL]].first().reset_index()
            df_result = df_result.merge(largest_calc, on='Stock Code', how='left').fillna(0)

            repo_base = df_pivot_full.rename(columns={'Local Code': 'Stock Code', 'Used Reverse Repo Qty': 'REPO_Base'})[['Stock Code', 'REPO_Base']]
            df_result = df_result.merge(repo_base, on='Stock Code', how='left').fillna(0)

            df_result['Total two Largerst'] = df_result[FIRST_LARGERST_COL] + df_result[SECOND_LARGERST_COL]
            df_result['Quantity Available'] = df_result['Quantity On Hand'] - df_result['Total two Largerst']
            df_result['Thirty Percent On Hand'] = 0.30 * df_result['Quantity On Hand']
            df_result['REPO'] = 0.10 * df_result['REPO_Base']
            df_result['Lendable Limit'] = np.minimum(df_result['Thirty Percent On Hand'], df_result['Quantity Available']) + df_result['REPO']
            df_result['Available Lendable Limit'] = df_result['Lendable Limit'] - df_result['Borrow Position']

            df_result_filtered = df_result[~df_result['Stock Code'].isin(STOCK_CODE_BLACKLIST)].copy()
            df_result_static = df_result_filtered[(df_result_filtered['Lendable Limit'] > 0) | (df_result_filtered['Available Lendable Limit'] > 0)].reindex(columns=FINAL_COLUMNS_LL)

            with pd.ExcelWriter(output_xlsx_buffer, engine='openpyxl') as writer:
                df_result_filtered.reindex(columns=FINAL_COLUMNS_LL).to_excel(writer, sheet_name=SHEET_RESULT_NAME_SOURCE, index=False)
                df_instr_old_raw.to_excel(writer, sheet_name=SHEET_INST_OLD, index=False, header=False)
            output_xlsx_buffer.seek(0)

            wb_template = load_workbook(template_file_data)
            ws = wb_template.active
            ws["B4"] = datetime.now().strftime('%d-%b-%y')

            f_body = Font(name='Roboto Condensed', size=9)
            align_center = Alignment(horizontal='center', vertical='center')
            align_left = Alignment(horizontal='left', vertical='center')
            border_thin = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            for r_idx, row in enumerate(df_result_static.itertuples(index=False), start=7):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    if c_idx >= 3:
                        try:
                            cell.value = int(round(float(value), 0))
                        except:
                            cell.value = value
                    else:
                        cell.value = value
                    cell.border = border_thin
                    cell.font = f_body
                    if c_idx == 1:
                        cell.alignment = align_center
                    elif c_idx == 2:
                        cell.alignment = align_left
                    else:
                        cell.alignment = align_center
                        cell.number_format = '#,##0'

            output_template_full = BytesIO()
            wb_template.save(output_template_full)
            output_template_full.seek(0)

            return output_xlsx_buffer, output_template_full, df_result_static

        except Exception as e:
            st.error(f"❌ Error Detail: {e}")
            return None, None, None

# ============================================================
# FUNGSI TEMPLATE EKSTERNAL
# ============================================================
def fill_simple_ll_template(df_result, template_buffer):
    wb = load_workbook(template_buffer)
    ws = wb.active
    ws["B4"] = datetime.now().strftime('%d-%b-%y')

    start_row = 7
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)

    f_body = Font(name='Roboto Condensed', size=9)
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for r_idx, row in enumerate(df_result.itertuples(index=False), start=start_row):
        c1 = ws.cell(row=r_idx, column=1, value=str(row[0]))
        c1.font, c1.border, c1.alignment = f_body, border_thin, align_center
        c2 = ws.cell(row=r_idx, column=2, value=str(row[1]))
        c2.font, c2.border, c2.alignment = f_body, border_thin, align_left
        val_c = 0
        try:
            val_c = int(round(float(row[2]), 0))
        except:
            pass
        c3 = ws.cell(row=r_idx, column=3, value=val_c)
        c3.font, c3.border, c3.alignment = f_body, border_thin, align_center
        c3.number_format = '#,##0'

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ─────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────
st.markdown("""
<div class="main-header">
  <h1>📉 Lendable Limit Calculation</h1>
  <p>Upload file input dan template untuk generate laporan LL</p>
</div>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────
# MAIN UI
# ─────────────────────────────────────────
def main():
    # Card: Upload File Input
    st.markdown('<div class="card"><div class="card-title">File Input</div>', unsafe_allow_html=True)
    cols = st.columns(3)
    u1 = cols[0].file_uploader('1. Instrument', type=['xlsx'])
    u2 = cols[1].file_uploader('2. Stock Position', type=['xlsx'])
    u3 = cols[2].file_uploader('3. BorrPosition', type=['xlsx'])
    st.markdown('</div>', unsafe_allow_html=True)

    # Card: Upload Template
    st.markdown('<div class="card"><div class="card-title">Template</div>', unsafe_allow_html=True)
    col_t = st.columns(2)
    t_full   = col_t[0].file_uploader('4. Template Full', type=['xlsx'])
    t_simple = col_t[1].file_uploader('5. Template External', type=['xlsx'])
    st.markdown('</div>', unsafe_allow_html=True)

    if all([u1, u2, u3, t_full, t_simple]):
        if st.button("▶ Jalankan Perhitungan LL"):
            files = {
                'Instrument.xlsx': u1,
                'Stock Position Detail.xlsx': u2,
                'BorrPosition.xlsx': u3
            }
            res = process_lendable_limit(files, BytesIO(t_full.getvalue()))

            if res[0]:
                xlsx_buf, full_buf, df_stat = res

                # Preview Result
                st.markdown('<div class="card"><div class="card-title">Preview Result</div>', unsafe_allow_html=True)
                st.dataframe(df_stat.style.apply(highlight_negative_ll, axis=1), use_container_width=True)
                st.markdown('</div>', unsafe_allow_html=True)

                # Proses External
                simple_df  = df_stat[['Stock Code', 'Stock Name', 'Available Lendable Limit']].copy()
                simple_buf = fill_simple_ll_template(simple_df, BytesIO(t_simple.getvalue()))

                # Download Buttons
                st.markdown('<div class="card"><div class="card-title">Unduh Hasil</div>', unsafe_allow_html=True)
                d_cols = st.columns(3)
                d_cols[0].download_button(
                    "⬇ Konsolidasi",
                    xlsx_buf,
                    "Konsolidasi.xlsx",
                    use_container_width=True,
                )
                d_cols[1].download_button(
                    "⬇ LL Lengkap",
                    full_buf,
                    f"Lendable Limit {datetime.now().strftime('%Y%m%d')}.xlsx",
                    use_container_width=True,
                )
                d_cols[2].download_button(
                    "⬇ LL Eksternal",
                    simple_buf,
                    f"Lendable Limit_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    use_container_width=True,
                )
                st.markdown('</div>', unsafe_allow_html=True)
    else:
        st.markdown("""
        <div class="card">
          <div class="card-title">Panduan</div>
          <p style="color:#8a8fa8;font-size:0.85rem;margin:0">
            Upload semua 5 file di atas, lalu klik <strong style="color:#5c7cfa">▶ Jalankan Perhitungan LL</strong> untuk memproses.
          </p>
        </div>
        """, unsafe_allow_html=True)

main()
