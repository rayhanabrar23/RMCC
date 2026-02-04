import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime

# ===============================================================
# 1. KONFIGURASI HALAMAN & LOGIN
# ===============================================================
st.set_page_config(page_title="HCCL Integrator Pro", layout="wide")

if "login_status" not in st.session_state or not st.session_state["login_status"]:
    st.session_state["login_status"] = True # Set True untuk testing

# ===============================================================
# 2. KONFIGURASI GLOBAL & LOGIKA PERHITUNGAN
# ===============================================================
COL_RMCC = 'CONCENTRATION LIMIT USULAN RMCC'
COL_LISTED = 'CONCENTRATION LIMIT TERKENA % LISTED SHARES'
COL_FF = 'CONCENTRATION LIMIT TERKENA % FREE FLOAT'
COL_PERHITUNGAN = 'CONCENTRATION LIMIT SESUAI PERHITUNGAN'
THRESHOLD_5M = 5_000_000_000
KODE_EFEK_KHUSUS = ['LPKR', 'MLPL', 'NOBU', 'PTPP', 'SILO']
OVERRIDE_MAPPING = {
    'LPKR': 10_000_000_000, 'MLPL': 10_000_000_000,
    'NOBU': 10_000_000_000, 'PTPP': 50_000_000_000, 'SILO': 10_000_000_000
}
TOLERANCE = 1e-6

# --- FUNGSI HELPER (DARI KODE KAMU) ---
def calc_concentration_limit_listed(row):
    try:
        if row['PERBANDINGAN DENGAN LISTED SHARES (Sesuai Perhitungan)'] >= 0.05:
            return 0.0499 * row['LISTED SHARES'] * row['CLOSING PRICE']
        return None
    except: return None

def calc_concentration_limit_ff(row):
    try:
        if row['PERBANDINGAN DENGAN FREE FLOAT (Sesuai Perhitungan)'] >= 0.20:
            return 0.1999 * row['FREE FLOAT (DALAM LEMBAR)'] * row['CLOSING PRICE']
        return None
    except: return None

def override_rmcc_limit(row):
    kode = row['KODE EFEK']
    nilai_rmcc = row[COL_RMCC]
    if kode in OVERRIDE_MAPPING:
        if nilai_rmcc == 0.0: return 0.0
        return min(nilai_rmcc, OVERRIDE_MAPPING[kode])
    return nilai_rmcc

def calculate_concentration_limit(df_source):
    df = df_source.copy()
    if 'KODE EFEK' not in df.columns:
        df = df.rename(columns={df.columns[0]: 'KODE EFEK'})
    
    df['KODE EFEK'] = df['KODE EFEK'].astype(str).str.strip()
    
    # 1. Logika Margin Baru
    df['SAHAM MARJIN BARU?'] = df['SAHAM MARJIN BARU?'].astype(str).str.upper().str.strip()
    df['CL_MARJIN_BARU'] = np.where(df['SAHAM MARJIN BARU?'] == 'YA', df[COL_PERHITUNGAN] * 0.50, df[COL_PERHITUNGAN])

    # 2. Listed & FF
    df[COL_LISTED] = df.apply(calc_concentration_limit_listed, axis=1)
    df[COL_FF] = df.apply(calc_concentration_limit_ff, axis=1)

    # 3. Min Value
    cols_to_min = ['CL_MARJIN_BARU', COL_LISTED, COL_FF, COL_PERHITUNGAN]
    df['MIN_VAL'] = df[cols_to_min].fillna(np.inf).min(axis=1)
    
    mask_nol = (df[cols_to_min].fillna(np.inf) < THRESHOLD_5M).any(axis=1)
    df[COL_RMCC] = np.where(mask_nol, 0.0, df['MIN_VAL'])
    
    # 4. Override & Haircut
    mask_not_zero = (df[COL_RMCC] != 0.0)
    df.loc[mask_not_zero, COL_RMCC] = df.loc[mask_not_zero].apply(override_rmcc_limit, axis=1).round(0)
    
    df['HAIRCUT PEI USULAN DIVISI'] = np.where(df['UMA'].notna(), df['HAIRCUT KPEI'], df['HAIRCUT PEI'])
    
    # Keterangan
    df['PERTIMBANGAN DIVISI (HAIRCUT)'] = df['UMA'].apply(lambda x: f"UMA BEI {x}" if pd.notna(x) else "Sesuai Metode")
    df['PERTIMBANGAN DIVISI (CONC LIMIT)'] = "Sesuai metode perhitungan" # Default
    
    return df

# ===============================================================
# 3. FUNGSI INJECTOR KE TEMPLATE (REVISI LOKASI)
# ===============================================================
def update_template_final(file_template, df_hasil):
    output = BytesIO()
    wb = load_workbook(file_template)
    
    # Mapping Sheet
    ws_hc = wb["HC"] if "HC" in wb.sheetnames else wb.active
    ws_conc = wb["CONC"] if "CONC" in wb.sheetnames else wb.active

    # Loop data hasil (i=0, row=data) -> Baris Excel mulai dari 5
    for i, row_data in enumerate(df_hasil.to_dict('records'), start=5):
        # --- SHEET HC ---
        ws_hc.cell(row=i, column=3, value=row_data.get('KODE EFEK'))                    # C5
        ws_hc.cell(row=i, column=18, value=row_data.get('HAIRCUT PEI USULAN DIVISI'))   # R5
        ws_hc.cell(row=i, column=20, value=row_data.get('PERTIMBANGAN DIVISI (HAIRCUT)')) # T5

        # --- SHEET CONC ---
        ws_conc.cell(row=i, column=3, value=row_data.get('KODE EFEK'))                  # C5
        ws_conc.cell(row=i, column=19, value=row_data.get(COL_RMCC))                    # S5
        ws_conc.cell(row=i, column=22, value=row_data.get('PERTIMBANGAN DIVISI (CONC LIMIT)')) # V5

    wb.save(output)
    output.seek(0)
    return output

# ===============================================================
# 4. ANTARMUKA STREAMLIT
# ===============================================================
def main():
    st.title("🛡️ HC & CL Processor to Template")
    st.info("Input: Raw Data -> Proses Koding -> Output: Template Excel Terisi")

    col1, col2 = st.columns(2)
    with col1:
        raw_file = st.file_uploader("Upload Raw Data (XLSX)", type=['xlsx'])
    with col2:
        tpl_file = st.file_uploader("Upload Template Target (XLSX)", type=['xlsx'])

    if raw_file and tpl_file:
        if st.button("🚀 Jalankan Proses", type="primary"):
            try:
                # 1. Hitung
                df_raw = pd.read_excel(raw_file)
                df_final = calculate_concentration_limit(df_raw)
                
                # 2. Inject ke Template
                tpl_file.seek(0)
                final_excel = update_template_final(tpl_file, df_final)
                
                st.success("✅ Selesai! Data telah dipindahkan ke Sheet HC (C, R, T) dan CONC (C, S, V).")
                
                # 3. Download
                st.download_button(
                    label="⬇️ Download File Hasil",
                    data=final_excel,
                    file_name=f"HCCL_Final_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"Error: {e}")

if __name__ == "__main__":
    main()
