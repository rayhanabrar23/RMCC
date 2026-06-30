import streamlit as st
import pandas as pd
import io
import copy
import base64
from datetime import date
from calendar import monthrange
import openpyxl
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Rekap SID - Isi Template Excel", layout="wide")
st.title("Rekap SID, Nama Partisipan, Nilai Pendanaan, Nilai Jaminan, Maturity, Status")
st.caption(
    "Versi ini langsung mengisi ke **template Excel asli** kamu (format, font, kolom hidden "
    "F/H/K, border, lebar kolom — semua persis sama). Hasilnya tinggal di-export ke PDF."
)
st.caption(
    "ℹ️ Setiap baris hasil = 1 kontrak unik (per SID). Nilai Jaminan otomatis di-SUM hanya untuk "
    "agunan dalam SID yang sama (misal basket saham margin) — kontrak berbeda dengan Nama "
    "Partisipan sama (misal beberapa kontrak REPO) tetap tampil sebagai baris terpisah."
)

def get_template_bytes():
    """Template di-embed sebagai base64 supaya tidak bergantung pada path file
    eksternal (penting untuk Streamlit multipage app, karena cwd bisa beda
    dengan folder file pages/...)."""
    return io.BytesIO(base64.b64decode(TEMPLATE_XLSX_B64))

# ----------------------------------------------------------
# SIDEBAR - mapping kolom
# ----------------------------------------------------------
st.sidebar.header("⚙️ Mapping Kolom (ubah jika perlu)")
st.sidebar.markdown("**File A01**")
col_a01_no_fas  = st.sidebar.number_input("A01 - posisi kolom No Fasilitas (SID)", min_value=0, value=2)
col_a01_nama    = st.sidebar.number_input("A01 - posisi kolom Nama Partisipan",    min_value=0, value=11)
col_a01_jaminan = st.sidebar.number_input("A01 - posisi kolom Nilai Jaminan",      min_value=0, value=15)

st.sidebar.markdown("**File F06**")
col_f06_sid       = st.sidebar.number_input("F06 - posisi kolom SID",                      min_value=0, value=1)
col_f06_cif       = st.sidebar.number_input("F06 - posisi kolom CIF",                      min_value=0, value=2)
col_f06_pendanaan = st.sidebar.number_input("F06 - posisi kolom Nilai Pendanaan",           min_value=0, value=9)
col_f06_maturity  = st.sidebar.number_input("F06 - posisi kolom Maturity",                  min_value=0, value=6)
col_f06_jenis     = st.sidebar.number_input("F06 - posisi kolom Kode Jenis Fasilitas Lain", min_value=0, value=3)

st.sidebar.markdown("**D02 (opsional, fallback nama)**")
col_d02_cif  = st.sidebar.number_input("D02 - posisi kolom CIF",  min_value=0, value=1)
col_d02_nama = st.sidebar.number_input("D02 - posisi kolom Nama", min_value=0, value=3)

st.sidebar.info("Posisi dihitung mulai dari 0. Kolom A=0, B=1, C=2, ... L=11, dst.")

# ----------------------------------------------------------
# Mapping jenis fasilitas -> label kolom B di template
# ----------------------------------------------------------
JENIS_TO_LABEL = {
    "005": "Pendanaan Transaksi REPO",
    "007": "Pendanaan Transaksi Marjin",
}
JENIS_RAW_LABEL = {
    "001": "Pendanaan Kredit Kelolaan",
    "002": "Pendanaan Tagihan Akseptasi",
    "003": "Pendanaan Kewajiban Kepada Pemerintah",
    "004": "Pendanaan Tagihan Transaksi Derivatif",
    "005": "Pendanaan Transaksi REPO",
    "006": "Pendanaan Tagihan Non Pembiayaan",
    "007": "Pendanaan Transaksi Marjin",
    "008": "Pendanaan Bai Al Musawamah (Salam)",
    "009": "Pendanaan Tagihan Subrogasi",
    "900": "Pendanaan Lainnya",
}

def label_jenis(code):
    code = (code or "").strip()
    return JENIS_TO_LABEL.get(code, JENIS_RAW_LABEL.get(code, f"Pendanaan Lainnya ({code})"))

# ----------------------------------------------------------
# HELPER PARSING (sama seperti versi sebelumnya)
# ----------------------------------------------------------
def read_pipe_file(uploaded_file):
    content = uploaded_file.read().decode("utf-8", errors="ignore")
    rows, header_cols = [], []
    for line in content.splitlines():
        line = line.rstrip("\r\n")
        if not line:
            continue
        if line.startswith("H|"):
            header_cols = line.split("|")
            continue
        rows.append(line.split("|"))
    return rows, header_cols

def detect_periode(header_cols):
    try:
        yr = int(header_cols[3]); mo = int(header_cols[4])
        last_day = monthrange(yr, mo)[1]
        return yr, mo, date(yr, mo, last_day)
    except Exception:
        return None, None, None

def safe_get(row, idx):
    return row[idx] if idx < len(row) else ""

def parse_number(value):
    if not value:
        return 0.0
    try:
        return float(value.replace(",", "").strip())
    except ValueError:
        return 0.0

def parse_date(value):
    value = (value or "").strip()
    try:
        return pd.to_datetime(value, format="%Y%m%d").date()
    except Exception:
        return None

# ----------------------------------------------------------
# UPLOAD MULTI-BATCH
# ----------------------------------------------------------
st.subheader("📁 Upload Batch Data")
st.caption(
    "Upload **lebih dari satu set file** sekaligus (misalnya 1 set untuk data Margin, 1 set lagi "
    "untuk data REPO Bonds). Urutan file A01 dan F06 akan dipasangkan berdasarkan urutan upload "
    "(file A01 ke-1 dipasangkan dengan file F06 ke-1, dst). Boleh upload >3 file sekaligus."
)

col1, col2, col3 = st.columns(3)
with col1:
    files_a01 = st.file_uploader("Upload semua file A01 (.txt)", type=["txt"], key="a01", accept_multiple_files=True)
with col2:
    files_f06 = st.file_uploader("Upload semua file F06 (.txt)", type=["txt"], key="f06", accept_multiple_files=True)
with col3:
    files_d02 = st.file_uploader("Upload file D02 (.txt) — opsional, fallback nama", type=["txt"], key="d02", accept_multiple_files=True)

st.markdown("---")
periode_override = st.checkbox("Override periode pelaporan manual (jika header F06 tidak terbaca / beda bulan antar file)")
if periode_override:
    c1, c2 = st.columns(2)
    with c1:
        yr_in = st.number_input("Tahun", min_value=2000, max_value=2100, value=date.today().year)
    with c2:
        mo_in = st.number_input("Bulan", min_value=1, max_value=12, value=date.today().month)

# ----------------------------------------------------------
# PROCESS
# ----------------------------------------------------------
if files_a01 and files_f06:
    if len(files_a01) != len(files_f06):
        st.error(f"❌ Jumlah file A01 ({len(files_a01)}) dan F06 ({len(files_f06)}) tidak sama. "
                 f"Setiap batch harus punya pasangan A01 + F06.")
        st.stop()

    try:
        # ---- D02 fallback (gabungan semua file D02) ----
        d02_by_cif = {}
        for f in (files_d02 or []):
            rows, _ = read_pipe_file(f)
            for r in rows:
                if safe_get(r, 0).strip() != 'D':
                    continue
                cif  = safe_get(r, col_d02_cif).strip()
                nama = safe_get(r, col_d02_nama).strip()
                if cif and nama:
                    d02_by_cif[cif] = nama

        all_results = []
        detected_periods = []

        for i, (fa01, ff06) in enumerate(zip(files_a01, files_f06), start=1):
            a01_rows, _ = read_pipe_file(fa01)
            f06_rows, f06_header = read_pipe_file(ff06)

            yr, mo, rdate = detect_periode(f06_header)
            if rdate:
                detected_periods.append((yr, mo))

            # A01 -> jaminan per no fasilitas (SUM dalam SID yang sama)
            a01_by_fas = {}
            for r in a01_rows:
                if safe_get(r, 0).strip() != 'D':
                    continue
                no_fas  = safe_get(r, col_a01_no_fas).strip()
                nama    = safe_get(r, col_a01_nama).strip()
                jaminan = parse_number(safe_get(r, col_a01_jaminan))
                if no_fas:
                    if no_fas not in a01_by_fas:
                        a01_by_fas[no_fas] = [nama, 0.0]
                    a01_by_fas[no_fas][1] += jaminan

            df_a01_fas = pd.DataFrame(
                [(k, v[0], v[1]) for k, v in a01_by_fas.items()],
                columns=["fas_key", "Nama Partisipan", "Nilai Jaminan"]
            ) if a01_by_fas else pd.DataFrame(columns=["fas_key", "Nama Partisipan", "Nilai Jaminan"])

            # F06
            f06_records = []
            for r in f06_rows:
                no_fas    = safe_get(r, col_f06_sid).strip()
                cif       = safe_get(r, col_f06_cif).strip()
                pendanaan = parse_number(safe_get(r, col_f06_pendanaan))
                maturity  = parse_date(safe_get(r, col_f06_maturity))
                jenis     = safe_get(r, col_f06_jenis).strip()
                if no_fas:
                    f06_records.append((no_fas, cif, pendanaan, maturity, jenis))

            df_f06 = pd.DataFrame(
                f06_records,
                columns=["SID", "CIF", "Nilai Pendanaan", "Maturity", "Kode Jenis"]
            )
            if df_f06.empty:
                st.warning(f"⚠️ Batch #{i}: file F06 ({ff06.name}) tidak menghasilkan baris data.")
                continue

            df_f06["Tipe Pendanaan"] = df_f06["Kode Jenis"].apply(label_jenis)

            result = pd.merge(df_f06, df_a01_fas, left_on="SID", right_on="fas_key", how="left")
            result = result.drop(columns=["fas_key"], errors="ignore")

            if d02_by_cif:
                result["Nama_D02"] = result["CIF"].map(d02_by_cif)
                result["Nama Partisipan"] = result["Nama Partisipan"].fillna(result["Nama_D02"])
                result = result.drop(columns=["Nama_D02"], errors="ignore")

            result["Nilai Jaminan"] = result["Nilai Jaminan"].fillna(0)
            result["Batch"] = f"{fa01.name} + {ff06.name}"
            all_results.append(result)

        if not all_results:
            st.error("❌ Tidak ada data valid dari batch manapun.")
            st.stop()

        final = pd.concat(all_results, ignore_index=True)

        # ---- Tentukan periode laporan ----
        if periode_override:
            report_year, report_month = int(yr_in), int(mo_in)
        elif detected_periods:
            report_year, report_month = detected_periods[0]
            if len(set(detected_periods)) > 1:
                st.warning(
                    "⚠️ Header F06 antar batch menunjukkan periode bulan/tahun yang berbeda-beda. "
                    "Memakai periode dari batch pertama. Centang 'Override periode pelaporan manual' "
                    "di atas jika ingin set manual."
                )
        else:
            st.error("❌ Gagal membaca periode dari header F06 di semua batch, dan override manual tidak dicentang.")
            st.stop()

        last_day = monthrange(report_year, report_month)[1]
        report_date = date(report_year, report_month, last_day)

        n_missing = final["Nama Partisipan"].isna().sum()
        if n_missing > 0:
            st.warning(f"⚠️ {n_missing} fasilitas tidak ditemukan nama nasabahnya di A01/D02.")

        final = final[["Tipe Pendanaan", "Nama Partisipan", "Nilai Pendanaan", "Nilai Jaminan", "Maturity", "Batch"]]
        final = final.sort_values(["Tipe Pendanaan", "Nama Partisipan"], na_position="last").reset_index(drop=True)

        st.success(
            f"✅ Berhasil! {len(final)} baris dari {len(all_results)} batch digabung. "
            f"Periode: {report_date.strftime('%d %B %Y')}"
        )

        st.subheader("📋 Preview Data Gabungan (sebelum dimasukkan ke template)")
        st.dataframe(
            final.style.format({"Nilai Pendanaan": "{:,.0f}", "Nilai Jaminan": "{:,.0f}"}),
            use_container_width=True,
        )

        # ----------------------------------------------------------
        # ISI KE TEMPLATE EXCEL ASLI
        # ----------------------------------------------------------
        wb = openpyxl.load_workbook(get_template_bytes())
        ws = wb.worksheets[0]

        TEMPLATE_FIRST_DATA_ROW = 5
        TEMPLATE_LAST_DATA_ROW  = 37  # jumlah baris siap pakai di template asli

        # Judul sheet & cell tanggal
        new_sheet_name = f"{report_month:02d}-{report_year}"
        try:
            ws.title = new_sheet_name
        except Exception:
            pass
        ws["B3"] = pd.Timestamp(report_date)

        n_rows = len(final)
        n_existing = TEMPLATE_LAST_DATA_ROW - TEMPLATE_FIRST_DATA_ROW + 1  # 33

        # Style "prototype" diambil dari baris pertama data template (row 5) -> dipakai utk baris baru jika perlu
        proto_row = TEMPLATE_FIRST_DATA_ROW

        def clone_row_style(src_row, dst_row, max_col=11):
            for c in range(1, max_col + 1):
                src_cell = ws.cell(row=src_row, column=c)
                dst_cell = ws.cell(row=dst_row, column=c)
                dst_cell.font = copy.copy(src_cell.font)
                dst_cell.border = copy.copy(src_cell.border)
                dst_cell.fill = copy.copy(src_cell.fill)
                dst_cell.alignment = copy.copy(src_cell.alignment)
                dst_cell.number_format = src_cell.number_format
                dst_cell.protection = copy.copy(src_cell.protection)
            ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height

        # Jika data lebih banyak dari kapasitas template, sisipkan baris baru sebelum baris terakhir+1
        if n_rows > n_existing:
            extra = n_rows - n_existing
            insert_at = TEMPLATE_LAST_DATA_ROW + 1
            ws.insert_rows(insert_at, amount=extra)
            for offset in range(extra):
                clone_row_style(TEMPLATE_LAST_DATA_ROW, insert_at + offset)

        last_row_used = TEMPLATE_FIRST_DATA_ROW + n_rows - 1

        for i, (_, row) in enumerate(final.iterrows()):
            r = TEMPLATE_FIRST_DATA_ROW + i

            # No urut
            if r == TEMPLATE_FIRST_DATA_ROW:
                ws.cell(row=r, column=1).value = 1
            else:
                ws.cell(row=r, column=1).value = f"=A{r-1}+1"

            ws.cell(row=r, column=2).value = row["Tipe Pendanaan"]          # B
            ws.cell(row=r, column=3).value = row["Nama Partisipan"]        # C
            ws.cell(row=r, column=4).value = row["Nilai Pendanaan"]        # D
            ws.cell(row=r, column=5).value = row["Nilai Jaminan"]          # E
            ws.cell(row=r, column=6).value = 0                              # F - manual (Net Cash Shortfall), default 0
            if pd.notna(row["Maturity"]):
                ws.cell(row=r, column=7).value = pd.Timestamp(row["Maturity"])  # G
            else:
                ws.cell(row=r, column=7).value = 0
            ws.cell(row=r, column=8).value = f"=IF(ISBLANK(K{r}),0,$B$3-K{r})"  # H
            ws.cell(row=r, column=9).value = (
                f'=IF(B{r}="Pendanaan Transaksi Marjin",IF(F{r}<=0,"LANCAR", "CEK JUMLAH HARI"),'
                f'IF(G{r}>$B$3,"LANCAR","CEK JUMLAH HARI"))'
            )  # I
            # K (Tanggal Macet) sengaja dikosongkan -> diisi manual oleh user jika ada kontrak macet

        # Bersihkan baris sisa template yang tidak terpakai (kalau data < kapasitas template)
        if n_rows < n_existing:
            for r in range(TEMPLATE_FIRST_DATA_ROW + n_rows, TEMPLATE_LAST_DATA_ROW + 1):
                for c in range(1, 12):
                    cell = ws.cell(row=r, column=c)
                    if c == 1:
                        cell.value = None
                    elif c in (8, 9):
                        cell.value = None
                    else:
                        cell.value = None

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        st.markdown("---")
        st.download_button(
            label="⬇️ Download Hasil (format template asli, siap di-PDF-kan)",
            data=buffer,
            file_name=f"Penilaian_Kualitas_Pendanaan_Transaksi_Efek_{report_month:02d}-{report_year}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        st.error(f"Terjadi error saat memproses file: {e}")
        st.exception(e)
else:
    st.info("⬆️ Silakan upload minimal satu pasang file A01 dan F06 untuk memulai (boleh upload banyak pasangan sekaligus).")


# ----------------------------------------------------------
# TEMPLATE EXCEL ASLI (di-embed base64 supaya tidak perlu file terpisah)
# ----------------------------------------------------------
TEMPLATE_XLSX_B64 = """
UEsDBBQABgAIAAAAIQB0NlqmegEAAIQFAAATAAgCW0NvbnRlbnRfVHlwZXNdLnhtbCCiBAIooAACAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACsVM1OAjEQvpv4DpteDVvwYIxh4YB6VBLwAWo7sA3d
tukMCG/vbEFiDEIIXLbZtvP9TGemP1w3rlhBQht8JXplVxTgdTDWzyvxMX3tPIoCSXmjXPBQiQ2gGA5ub/rTTQQsONpjJWqi+CQl
6hoahWWI4PlkFlKjiH/TXEalF2oO8r7bfZA6eAJPHWoxxKD/DDO1dFS8rHl7q+TTelGMtvdaqkqoGJ3VilioXHnzh6QTZjOrwQS9
bBi6xJhAGawBqHFlTJYZ0wSI2BgKeZAzgcPzSHeuSo7MwrC2Ee/Y+j8M7cn/rnZx7/wcyRooxirRm2rYu1w7+RXS4jOERXkc5NzU
5BSVjbL+R/cR/nwZZV56VxbS+svAJ3QQ1xjI/L1cQoY5QYi0cYDXTnsGPcVcqwRmQly986sL+I19QodWTo9qLpErJ2GPe4yfW3qc
QkSeGgnOF/DTom10JzIQJLKwb9JDxb5n5JFzsWNoZ5oBc4Bb5hk6+AYAAP//AwBQSwMEFAAGAAgAAAAhALVVMCP0AAAATAIAAAsA
CAJfcmVscy8ucmVscyCiBAIooAACAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAACskk1PwzAMhu9I/IfI99XdkBBCS3dBSLshVH6ASdwPtY2jJBvdvyccEFQagwNHf71+/Mrb3TyN6sgh9uI0rIsSFDsj
tnethpf6cXUHKiZylkZxrOHEEXbV9dX2mUdKeSh2vY8qq7iooUvJ3yNG0/FEsRDPLlcaCROlHIYWPZmBWsZNWd5i+K4B1UJT7a2G
sLc3oOqTz5t/15am6Q0/iDlM7NKZFchzYmfZrnzIbCH1+RpVU2g5abBinnI6InlfZGzA80SbvxP9fC1OnMhSIjQS+DLPR8cloPV/
WrQ08cudecQ3CcOryPDJgosfqN4BAAD//wMAUEsDBBQABgAIAAAAIQBAiXbYpQMAABwJAAAPAAAAeGwvd29ya2Jvb2sueG1srFZd
b6M4FH1faf8D4t3FBvOppqMQQFupHVWdTPtSaeSCU1ABZ43TpKrmv881gaRpVqtsZ6PExr7m+Fzfc69z/mXT1MYLl10l2olJzrBp
8DYXRdU+Tczv8wwFptEp1hasFi2fmK+8M79c/PnH+VrI50chng0AaLuJWSq1jCyry0vesO5MLHkLloWQDVMwlE9Wt5ScFV3JuWpq
y8bYsxpWteYWIZKnYIjFosp5IvJVw1u1BZG8Zgrod2W17Ea0Jj8FrmHyebVEuWiWAPFY1ZV67UFNo8mjy6dWSPZYg9sb4hobCV8P
fgRDY487geloq6bKpejEQp0BtLUlfeQ/wRYhB0ewOT6D05CoJflLpWO4YyW9T7LydljeHozg30YjIK1eKxEc3ifR3B0327w4X1Q1
v9tK12DL5VfW6EjVplGzTqVFpXgxMX0YijU/mJCrZbyqarDa1HGoaV3s5HwjjYIv2KpWcxDyCA+Z4Xmh7eqVIIxprbhsmeIz0SrQ
4eDX72qux56VAhRu3PK/V5XkkFigL/AVWpZH7LG7Yao0VrKemLPo4XsH7j9I9lqy9ow9SiYfErFuawF59vBOoOw4G/6DRFmu/bbA
8S257fPHQwCOMhpleKOkAc+XyRWE4ht7gcBA+Ishby/h5IMfb840ISSZYhQEPkF05mMUpnaG7DTDXuw7CYmdn+CF9KJcsJUqh2Br
zIlJIbJHpmu2GS0ER6uq2O//hocP0v2HZrT91J7qsnZX8XW3l4UeGpv7qi3EemIigqEsvh4O173xvipUCbpybBfSZzv3F6+eSmBM
bKonQf6a2cR8i90wI9NpjGwPx4i6cYzCMPAQjlPbzVInoYHbM7LeUeoLKFDre6PtRY8JsrFNoVTr6qqPlzimISO9i7wsSB+/8cWc
1TnIXHf9ypBgO9Qr+EZddarvQWEVECQUT30cUoRTx0U0CG0UUMdGM5rYqeunSRoDwTGt/4dC2As9Gu8WzbJkUs0ly5/hRrrli5h1
oKWtQ8D3PdnYDWLsAEWakQxREmIUxx5FbpI5rk+SWepme7La/cUny1Bg9W9zplaQojo7+3Gk22yY3U0uthNDpA7SLrpN9LkPb//b
wm/gfc1PXJzdnbhw9vV6fn3i2qt0/uM+64X0j95aH6KREBpiJ50ix5lRRP3MR0GGXeRQn85cGqcE+/to1Ov85XPBsKk1ymX2/hof
aoUOjgaPhv84RsfVYDqQkabfi3+HdvELAAD//wMAUEsDBBQABgAIAAAAIQCSB5TsBAEAAD8DAAAaAAgBeGwvX3JlbHMvd29ya2Jv
b2sueG1sLnJlbHMgogQBKKAAAQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACskstqxDAMRfeF/oPRvnEyfVCGcWbRUpht
m36AcJQ4TGIHW33k72tSOsnAkG6yMUjC9x6Ju9t/d634JB8aZxVkSQqCrHZlY2sF78XLzSOIwGhLbJ0lBQMF2OfXV7tXapHjp2Ca
PoioYoMCw9xvpQzaUIchcT3ZOKmc75Bj6WvZoz5iTXKTpg/SzzUgP9MUh1KBP5S3IIqhj87/a7uqajQ9O/3RkeULFjLw0MYFRIG+
JlbwWyeREeRl+82a9hzPQpP7WMrxzZYYsjUZvpw/BkPEE8epFeQ4WYS5XxNGY6ufDDZ2gjm1li5yt2ooDHoq39jHzM+zMW//wciz
2Oc/AAAA//8DAFBLAwQUAAYACAAAACEA8BeDEL8MAAA2RQAAGAAAAHhsL3dvcmtzaGVldHMvc2hlZXQxLnhtbJyVXW/aMBSG7yft
P1i+J9+QgEirFspWbRdVu49r4zjEahJntinQaf99x05CmTqxrBJwTJz3Oa/PsZP55b4q0ROTios6xb7jYcRqKjJeb1L89ctqlGCk
NKkzUoqapfjAFL68eP9uvhPyURWMaQSEWqW40LqZua6iBauIckTDapjJhayIhr9y46pGMpJZUVW6gedN3IrwGreEmRzCEHnOKVsK
uq1YrVuIZCXR4F8VvFE9raJDcBWRj9tmREXVAGLNS64PFopRRWe3m1pIsi5h3Xs/IhTtJXwC+IZ9Gnv9VaaKUymUyLUDZLf1/Hr5
U3fqEnokvV7/IIwfuZI9cdPAF1TwNkv++MgKXmDhG2GTI8yUS862PEvxz5UXXS3iq+vReJx4oyjyb0bX8U0ymoZBECTeylsul7/w
xTzj0GGzKiRZnuIrf/YpjLF7Mbcb6BtnO3UyRpqsH1jJqGaQxMfI7M+1EI/mxlu45AGyITVD+4cGupxiWOChG0YYadF8ZrlesLJM
8WKMEaGaP7E7UKR4LbQW1T3fFNqeBg3XcimeWW392LTGaKvpGAFA1A/rfQFjY/x4ozGSYsjZMf9T2Poxfnufg3P+sZTTtB98KFrn
14zBr3us9Om4r/rKnuw7iTKWk22p78XuIzMlgvJHDpTUHo1ZdlgyReGsQguc0GCpKIEBv6ji5pkDR43s257xTBcpjjFaM6VXpksw
S7cKqv+9nbPGjlroodVC3HXzUydJJrGfxFD+M0rYkVYJsVMGoTMeRxMrHJgeFmkhEPv0vhOGgRf6pvtn0sOsVULslcGJcmD+SUeB
2FOgdgXPMtYW9owDuNE6gNhrATMwMbwUrBji3+wPczDtIBB7iOfA8yDyJv+ontmq7Z4xB71rfODEcRxGoWn8WQOu3YC/AQAA//8A
AAD//6xbbW/bNhD+K4YwFC2WNuKLJDtNDFixLXtph6HbfoDRJmu3IRuSoNv+/UiJp5BH8kjZ7aeUfHiU7rnjy6Pz5ePn29un9eHp
sLx8+Ouf2cNVwYrZ49+H+0f11wVT//n8pP6aF7N/mTx8vPj03/r28ePtvWos34iqWF5+1INWepTCiWKmeh5V89cll5fnX5eX5x8N
ph0x56bl2mtZey0br2XrtXRey85r2dst5+pdxxfmR72wHqWdBe/S4oZr3LDGDRvcsMUNHW7Y4Ya91eC8lWLCoTFE38ieBl8VlU1e
hcgbIHXPrKwE527/9dDfjP5Y44YNbtjihg437HDD3mpw3lYGOOTzN5GwHd9bD7sqFGoM2hK9to9g6MV9BHLNOmBj4RrZBIygJ9n6
EIESrAtAhDvRLjAReqF9AILe6GaAqNAb/Sae38ghRsVUfhhq8FWx6GMMPVU79LHSjlH0ctcBjKhdB6wNhvWTIB9vqM4t1dmZTh4y
u7NG3vUu+3x4uP1UzB5u766KXXWxE4168y9qSS2W++3L/c/tu9WPNy9vqldn5dl37Xfitfrz8vxOr6rokffgs97u04Oa/07baKur
4qfb+0+H+8PhfvbLg1rQD388fpm9Pzz8/uW+OFOQbfXiz6e3V+VZoWa7Xn0ozmbF9eZm9sOv79+tdrPd6sO+eKWBXfXit6e3+jme
oT4SHnAwNnrdiYZ6SjRo8BANd8tV9b0KiN4BKBbbAUYHRgCDYmdtIENczOdMSs5qXqLw2TgwVsum4XNesjnOVgf3ulrwas6lQKjO
oMJRY5lwomYIFLXv+MEALrOCwYu3fX2x53odVfGm3lbHSp2OlTo3VupvFSs6JexzCLmBabCJFe+FV83FChJMeXpVj6GEYqAdrNCh
FMCgZXhtIOElhurcUp2d6QwHizUyN1jAZ5FgGeKjDzMqpdXmmU+TBgdo0lNxfXL4usRnxmEETUkAg04vawMJU0J1bqnOznSGKbFG
5lIC/jmNksUUSjSYpgSfBIcRNCUBDN6IDSRMCdW5pTo70xmmxBqZSwn45zRKtLfy06RH06Qgb7ZmCM1KCNSg8xFgwryQvVuyt4Pe
MDX22FxuRj+dSA6+8pJbjb4TJzIG+bQ1QxLkmAuqfcKdY3IMZiCHlWWzKBlvKnQm2sCEA46XYiGFEGWFTykuDp9OoDdCmPUs2YSB
7yYT1qsMN+aJhN4rnIMlw5d4msLh0qwvGt6Tm30Ieb7tJ1AHJvL+EQKhm94aMJH8Mtf5YO+WHNtBb4Quy3I2XeCnE/NrkhahNaRE
fiGftmZIgpzBrgNiKOTXYCnCjjERYYfq7cByhB1rbDY74Khjk2kYL/1kwmoKnUyDGEAlE3Zzy4YxCcICIIZkgTVYihBmTEQIo3o7
sBwhzBqbTRh46ljChvGB1W+SysKeZZbY6ofd3JoxCcIC+gvDYhhYGiiRTanExAqtthsXxETTyIXa7dzdcOuivN2LlGTssdn8BbQW
+7YZvzGZ3WsYH0i4SboIs4QRPL3ZvbDXWzMmwV9AJmFYJwFLkYRzdA+PMKq3A8uRhDtCDhk9deL+ZUsR6j5AL4cRKcK642KftixH
eAiBGJYeABRhhxQfyLEd9EbYOUJ/MBaVQ7WKPEiXeenkHv1sBSLJTlqBwD5t9ccv/amHPvoFQAyrEGApwg6pQ5BjO+iNsHOEFGEs
nsyOLUYk2UmLEdinLcuRI0IghgUJAEXYISUJcmwHvRF2jlAljMVT2eG2LJFipwfTqgT2aWvG0LkTAjGsSwAozA7ZuyV7O+gNs2OP
zT0ojJ46bWXjti6RZCctS2Cftv0EqZUtBGJYmABQhB1HtsCnAnJsB70Rdo4QIYzFk3PHlhyS7KQVB+zTlpsbO7nvhEAC7zsAirBD
ag7k2A56I+wcoTkYiyezY2sOSXbSkgPDmkP/CS2ZOwHNQXgrG6UbbGCe8BWW7O2gN8LOEZqDsXgyO7bCkGTHEhiWKy6evwTjYg0e
UAwELlQJgbjHiSMNzOel+YfKNsDWQA6rGlnKHoluqoALEtFBrzTFNcoEqtqw58nehBJqQ198sOfyYg8fR8XwNZjL9OdgLnO/Byvk
Nyoe4LbKkQyatMjBcYlJPwFKaT98QoUoXvg4tSKs4rH4cXCiboTktcbi+KG0jA6em4ifWAUKUUtgrJKp3ktRX5fU92FuaxtJ1tLS
Bi48a/sJkqyFKkCw+AuWhmwWQyb7bGxcXLOQi2ZRhVijyjs6sAKsCYm27Z09T3bWJ0pANOE5rE2RPHha8uC41MKMcS4Gfq6FdBHv
6OmIGmOm4RzawIwDu1Vdc7Ws62TDuUbVWXRgZWStQQeDnT1PNmsZtRg5rE2RQnhaCvEqeM2YBGshvcRjzRE7CNYcHMEaVYrRwXOP
rM1xXSQgdHRks5ZRrpHDmr7/Q6VTcoUcxALqawvHBRvcCAz2RcLPtQCIY4kELA05xMuGi7KRjUQ76cbF1aypFTC0QppJI+ci0zuy
tvByzSAmsQYeJC7nGayJKdJJD6alE+zr1oyhcy0E8lgDkDmlslr0u5W3Qro4UVVlUy8CrAEuzBr0xlmz58nNtdGDJ7I2RVIRaUkF
XxxaMybBWqDUQ+AVEixBrsEREtcsuzghxWJRM+bdIbaAi7BmHul5X0Nn2p09TzZrGfUeObk2RWoRaamF4+oOMybBWkCPEfh7NFgy
p4yRNI81R3RhrJlXpZDqGyc+joDBCG3GzPPRH73azn6gbNoy6j5yaJuiwYi0BsOxBmPGJGgLaDDY0hosmaM/F0yd6WWDdtKNC2OM
m4XUI83MGSHN9BrS6hLXR+zsebJJyygHySFtijQj0rUfuOa9NWMSpIUKRHApFVgy+9pcFROwuq4atHJtXFwt+YLVjWYOnfwBF2HN
PBLB2hE1IaMHT9zXpmgj6md7qQorvK61ZkyCtZA24rHmaB5SSLkoJVPJhrQ1mNGspJKxuqr1KRKzRmojYIVg7QhtZPTgiaxN0UZE
WhvBh/rWjEmwFtBG/AXSKd1grFSqh84i/EM2mNHkZBldIUltBKwQrB1RDjJ68ETWpmgjvbZKn/wF1kbMmARrIW3EyzVHGxFq0Wsa
dWnDmeagKjkPr46kLgLPTDB2RInI6L0jGTt//nnz/wAAAP//AAAA//90kEFuwjAQRa9i+QDFdtyAURIJsWJRseAEppkkFonHmgyq
1NPjiKqwgO38/2f+m2oC6mEP4ziLb7xGrqWWTfU/FQRdLXd6e9By1VSrh72pEoXIx8QB4ywGpPCLkf24h8hA0C6bciT5Hr489SGb
RujyAfVhitIYZdfGWuWsdVoKCv3wTmNMS2ptN6rQ1pWqNM4ZXUhxRmac3ogD+BZoEZ9T5tNtrBQdYm75WvxrfQK+JpG5MpBfKGuZ
kJh84Nx3GzIhHdr7X36QLvMAwM0NAAD//wMAUEsDBBQABgAIAAAAIQDBFxC+TgcAAMYgAAATAAAAeGwvdGhlbWUvdGhlbWUxLnht
bOxZzYsbNxS/F/o/DHN3/DXjjyXe4M9sk90kZJ2UHLW27FFWMzKSvBsTAiU59VIopKWXQm89lNJAAw299I8JJLTpH9EnzdgjreUk
m2xKWnYNi0f+vaen955+evN08dK9mHpHmAvCkpZfvlDyPZyM2Jgk05Z/azgoNHxPSJSMEWUJbvkLLPxL259+chFtyQjH2AP5RGyh
lh9JOdsqFsUIhpG4wGY4gd8mjMdIwiOfFsccHYPemBYrpVKtGCOS+F6CYlB7fTIhI+wNlUp/e6m8T+ExkUINjCjfV6qxJaGx48Oy
QoiF6FLuHSHa8mGeMTse4nvS9ygSEn5o+SX95xe3LxbRViZE5QZZQ26g/zK5TGB8WNFz8unBatIgCINae6VfA6hcx/Xr/Vq/ttKn
AWg0gpWmttg665VukGENUPrVobtX71XLFt7QX12zuR2qj4XXoFR/sIYfDLrgRQuvQSk+XMOHnWanZ+vXoBRfW8PXS+1eULf0a1BE
SXK4hi6FtWp3udoVZMLojhPeDINBvZIpz1GQDavsUlNMWCI35VqM7jI+AIACUiRJ4snFDE/QCLK4iyg54MTbJdMIEm+GEiZguFQp
DUpV+K8+gf6mI4q2MDKklV1giVgbUvZ4YsTJTLb8K6DVNyAvnj17/vDp84e/PX/06PnDX7K5tSpLbgclU1Pu1Y9f//39F95fv/7w
6vE36dQn8cLEv/z5y5e///E69bDi3BUvvn3y8umTF9999edPjx3a2xwdmPAhibHwruFj7yaLYYEO+/EBP53EMELEkkAR6Hao7svI
Al5bIOrCdbDtwtscWMYFvDy/a9m6H/G5JI6Zr0axBdxjjHYYdzrgqprL8PBwnkzdk/O5ibuJ0JFr7i5KrAD35zOgV+JS2Y2wZeYN
ihKJpjjB0lO/sUOMHau7Q4jl1z0y4kywifTuEK+DiNMlQ3JgJVIutENiiMvCZSCE2vLN3m2vw6hr1T18ZCNhWyDqMH6IqeXGy2gu
UexSOUQxNR2+i2TkMnJ/wUcmri8kRHqKKfP6YyyES+Y6h/UaQb8KDOMO+x5dxDaSS3Lo0rmLGDORPXbYjVA8c9pMksjEfiYOIUWR
d4NJF3yP2TtEPUMcULIx3LcJtsL9ZiK4BeRqmpQniPplzh2xvIyZvR8XdIKwi2XaPLbYtc2JMzs686mV2rsYU3SMxhh7tz5zWNBh
M8vnudFXImCVHexKrCvIzlX1nGABZZKqa9YpcpcIK2X38ZRtsGdvcYJ4FiiJEd+k+RpE3UpdOOWcVHqdjg5N4DUC5R/ki9Mp1wXo
MJK7v0nrjQhZZ5d6Fu58XXArfm+zx2Bf3j3tvgQZfGoZIPa39s0QUWuCPGGGCAoMF92CiBX+XESdq1ps7pSb2Js2DwMURla9E5Pk
jcXPibIn/HfKHncBcwYFj1vx+5Q6myhl50SBswn3Hyxremie3MBwkqxz1nlVc17V+P/7qmbTXj6vZc5rmfNaxvX29UFqmbx8gcom
7/Lonk+8seUzIZTuywXFu0J3fQS80YwHMKjbUbonuWoBziL4mjWYLNyUIy3jcSY/JzLaj9AMWkNl3cCcikz1VHgzJqBjpId1KxWf
0K37TvN4j43TTme5rLqaqQsFkvl4KVyNQ5dKpuhaPe/erdTrfuhUd1mXBijZ0xhhTGYbUXUYUV8OQhReZ4Re2ZlY0XRY0VDql6Fa
RnHlCjBtFRV45fbgRb3lh0HaQYZmHJTnYxWntJm8jK4KzplGepMzqZkBUGIvMyCPdFPZunF5anVpqr1FpC0jjHSzjTDSMIIX4Sw7
zZb7Wca6mYfUMk+5YrkbcjPqjQ8Ra0UiJ7iBJiZT0MQ7bvm1agi3KiM0a/kT6BjD13gGuSPUWxeiU7h2GUmebvh3YZYZF7KHRJQ6
XJNOygYxkZh7lMQtXy1/lQ000RyibStXgBA+WuOaQCsfm3EQdDvIeDLBI2mG3RhRnk4fgeFTrnD+qsXfHawk2RzCvR+Nj70DOuc3
EaRYWC8rB46JgIuDcurNMYGbsBWR5fl34mDKaNe8itI5lI4jOotQdqKYZJ7CNYmuzNFPKx8YT9mawaHrLjyYqgP2vU/dNx/VynMG
aeZnpsUq6tR0k+mHO+QNq/JD1LIqpW79Ti1yrmsuuQ4S1XlKvOHUfYsDwTAtn8wyTVm8TsOKs7NR27QzLAgMT9Q2+G11Rjg98a4n
P8idzFp1QCzrSp34+srcvNVmB3eBPHpwfzinUuhQQm+XIyj60hvIlDZgi9yTWY0I37w5Jy3/filsB91K2C2UGmG/EFSDUqERtquF
dhhWy/2wXOp1Kg/gYJFRXA7T6/oBXGHQRXZpr8fXLu7j5S3NhRGLi0xfzBe14frivlzZfHHvESCd+7XKoFltdmqFZrU9KAS9TqPQ
7NY6hV6tW+8Net2w0Rw88L0jDQ7a1W5Q6zcKtXK3WwhqJWV+o1moB5VKO6i3G/2g/SArY2DlKX1kvgD3aru2/wEAAP//AwBQSwME
FAAGAAgAAAAhAPYUe8faBAAALhIAAA0AAAB4bC9zdHlsZXMueG1s1Fhtb9s2EP4+YP9BUIZ9GKboxZJrO7bT2omBAl1RLBkwYCkC
WqJsohTpUXRqd9h/35GULNqOYycBhs0fLJHi3T3P3fGOUv9yVVDnAYuScDZww/PAdTBLeUbYbOD+djvxOq5TSsQyRDnDA3eNS/dy
+P13/VKuKb6ZYywdUMHKgTuXctHz/TKd4wKV53yBGTzJuSiQhKGY+eVCYJSVSqigfhQEbb9AhLlGQ69IT1FSIPFlufBSXiyQJFNC
iVxrXa5TpL33M8YFmlKAugpjlDqrsC0iZyVqI3p2z05BUsFLnstz0OvzPCcp3ofb9bs+ShtNoPllmsLED6It7ivxQk2xL/ADUeFz
h322LCaFLJ2UL5kcuPFmyjFP3mcw2XIdE5Qxz8BN995PztnPZ2fBeRDcexd320P19Mc/l1xeeOZyeQmL7r23957r1wYt7WE7PqDe
1q1VPFdxsq34jx+8OAo/Z9mdU8DvzlnD7+LtAVTtPeGw1Qm6n420LexXThz2c84aX76BsKuE6n1h/CubqEewX8DBatWwX35zHhCF
mVABSDnlwpGwEcDBeoahApsVY0TJVBC1LEcFoWszHakJvXeqdQWBTFaTvrHw79qZKjQ1p/gYp1/5lEvujDnLMCtxtgd7S92+iwId
tY2LnqWu+2pwNc/Xa7Jp7ms7naUOeQkJQijd7OaWSjaYGPah8Eks2AQGTnV/u15AqjGo0cb3et2R1TOB1mGUnC5QckoyhWI23kvw
aTVHIAVWGOoMFAKVvBZWlcoal74AvSkXGbSeulxFoNpMDfsU5xLEBZnN1VXyBfxDkkkoz8N+RtCMM0SVgVrCloSWBd1p4Mo5dJd6
O+4iUyYqCyet11g0lJOWA+Qa8UnrDbnHuVUkwWUppvRGkfs93/LbKreKMDRxVS1UtVe34PPq1vjIDMB3W0KmLxip8KCUgxYLulbl
T+s2IzDQjEY6qM34HSUzVmBb4JPgEqdSHzn0pvBtWoakxS9MIIbPJ+is8seZWu6JDhOtpW3G4BfN2OKkWgBUdEPRmXNBvoFzVCtI
gTM2FXyVHyag2nEVIUDTRAhsPQVBbX7VWwyg2ukvAaiOf5KkFmTnq0CLW7yCGOta7e/hr9r8sRyrGNSoPi6LKRYTfSJssJtk0oa2
gvw8GxBUHZpdG7tpbgVfnVQe3xsHPL+rK2zrY4XxwsnanvbCMxJMFTBdY59KL3VqehVJy2GK7f8uVS38h3xxZKu9ZnOpDvWS8JwM
aT8n9QHc5ORRwqftzN0a81TrONlkve2PKf8PVOuwrd89TvTpoUr0WCNpUksliuqC0PesLr/V4zdN0lFvEwN3zIsC1V0CIjldEioJ
U02vpQ/F9VmhWv9RFV5qtRVLYKcNA4Zs1Zww9FOpXqv12WODCqhmOEdLKm83Dwduc/8LzsiygLZWrfpEHrjUKgZuc/9BHfLCtoIM
PedDCacyuDpLQQbuX9ejN92r60nkdYJRx4tbOPG6yejKS+Lx6Opq0g2iYPy39XL/ild7/S0CGnUY90oKHwBERbYCf9PMDVxrYODr
QgywbezdqB28S8LAm7SC0IvbqON12q3EmyRhdNWOR9fJJLGwJy/8BBD4YWg+JijwSU+SAlPC6ljVEbJnIUgwfIKEX0fCbz70DP8B
AAD//wMAUEsDBBQABgAIAAAAIQDGKQM0nQIAAKsGAAAUAAAAeGwvc2hhcmVkU3RyaW5ncy54bWyMVU1v2kAQvVfqfxj5lB4a06hq
qwiIFjCwwV4se4mU4wBb2Nhe0/1oy7/vOqmUCOOWoz3jNzPvzRv3735XJfwU2shaDYJP170AhNrUW6l2g2DFpx+/BWAsqi2WtRKD
4ChMcDd8/65vjAX/rTKDYG/t4TYMzWYvKjTX9UEoH/le6wqtf9S70By0wK3ZC2GrMrzp9b6EFUoVwKZ2yg6Cr76KU/KHE+OXF597
wbBv5LBvh2x53Q/tsB82jy+vOE0jSCM2IYwQBjwjLCeLnEI0jRanyYwkBFKScZrTlLDTcMLGkIvCaWnRnAbj2joDpJkeVXfao0Nl
8TUOVG09W0biKeBiBA9Yrp0+dqNxLSuxw/05PODrokXHYkYvqZ27Ta27y44yChNUqEVh3szS4guPa1TFJQXnQnnWUmflXqPsrpx7
OhCm+IT/as9zjDDGg9ep7MaK9FYWe4SolLY7K0W1QZiV9foZa9OoL0VL/mw+umRO8oRy/SaRnBGe0ZjQ151tLelz+J4klLVX9H6V
xGQOc5LR1j6tSEw5ybuRU+G3V6EXgmtUBgsjIUH9JNUpljdU04U31KKFeuIxSPkb/zWuA8omSxbllLTMimq38zTftupxGOHRwUjo
Uvr+Eiwqp89kJd4Nnt4SNRpYWaxatvLtND5+kM0ZAybsr1oX5qxZiBEHyF2J/lZ5S6Nf5u05PL+2GmFUayVqIDtdQyz8IdTyTION
bYC7au32kApdoGkjnpEhi9Jly16ErzLKH2FCeARNBlzdsAnE0exDi9klJzGwiMOY5HPI58uMT0kcw1VCshllrQ+eifyvhUaukgdn
hR+/8zBmopFsIy6+UV6hv3ZLxFbipq4adVoTETab+ZkSMo74azD0P5vhHwAAAP//AwBQSwMEFAAGAAgAAAAhADttMkvBAAAAQgEA
ACMAAAB4bC93b3Jrc2hlZXRzL19yZWxzL3NoZWV0MS54bWwucmVsc4SPwYrCMBRF9wP+Q3h7k9aFDENTNyK4VecDYvraBtuXkPcU
/XuzHGXA5eVwz+U2m/s8qRtmDpEs1LoCheRjF2iw8HvaLb9BsTjq3BQJLTyQYdMuvpoDTk5KiceQWBULsYVRJP0Yw37E2bGOCamQ
PubZSYl5MMn5ixvQrKpqbfJfB7QvTrXvLOR9V4M6PVJZ/uyOfR88bqO/zkjyz4RJOZBgPqJIOchF7fKAYkHrd/aea30OBKZtzMvz
9gkAAP//AwBQSwMEFAAGAAgAAAAhAJ4nwya7AQAANBUAACcAAAB4bC9wcmludGVyU2V0dGluZ3MvcHJpbnRlclNldHRpbmdzMS5i
aW7slM9K41AUxr82oxY3KghuXIhLabGl8c9SaaJWElOSVLoSio0Q0KSkEZkRFzLreQMfpo/QB3DlwoWID+Bm5ruxogxFKrgRzg3n
nj/360nyIz02QhwjQYwe7QQpltBgHiLK4pRVVTGwg1Er90ObvEVrXivloK6H6bjQoZ9BK5+nb+U17hYCdku5JyO7fK6YG8qVz9OU
/8u1W/f0952M+kFzGQMUteLc3dEf76O7TGSHq1mvL3hEafENCbx+V+M8+oAiz/b3lXYWfVyijE3o/JeUUeG+jRJMrKPKWolmYINX
iZoq6yajMnOdeYW+xqyKtSy7YkfX9AzLQjMKk6Cnoka7GyRe+CuAZfq+6cJJwiBK22kYR2g4ru9u1324QS8+Pc9qDJ2uiiqoxadx
Ysed4CUa/XbFOeBQN+xXBjfT3eVFSh9pGu055xT0+wv799PU3kJ/7Vq9vzU8Q+Gtp9KqfGXoVb5FO1T5LMgh5rw5xxlngZowTc4d
NRUaaDPq4YLnCToU/690eBaNqa2xx0902d/jL9T91ERLWZMlBISAEBACQkAICAEhIASEgBAQAkJACAiBcQj8AwAA//8DAFBLAwQU
AAYACAAAACEAi+6t578BAACKCAAAEAAAAHhsL2NhbGNDaGFpbi54bWx0ls1OwzAQhO9IvEPkO01tQ/lR06pCQuwdHiBKTRMpcao4
QvD2GKF62x18irLTdWfsz9uut19DX3y6KXSjr5ReLFXhfDPuO3+o1Pvby82DKsJc+33dj95V6tsFtd1cX62bum+e27rzRVzBh0q1
83x8KsvQtG6ow2I8Oh+Vj3Ea6jm+TocyHCdX70Pr3Dz0pVkuV+UQF1CbdVNMlXq196roogmrij4+VJmE1Ung0h2WbrEUl/pbkRsN
ljSW4j6IRvOIpbg58lMpQ/pGg+4Nujfo3qB7g+4NujfoXqN7je41utfoXqN7je5/j1Fsjkb3Gt1rdI/m0TtaR+dgnJi503mRhT6y
/zRCYrKQmCwkJguJyUJiQtrIQGYyEJqQNkLaCGkjpI2QNkLaCGkjpI2QNkLaCGkjpI2QNkLaCGkjpI2QNgLaCHceNx6o2WGFB1wc
mDzgdmn94rKenIg6m5ZCgksKCUQpJGilkACXQroMUkiBpZDLzVDIjlxyhk10MJhSyCVn4GVHLjlfJNmRS84XVHbkjpwvvujgISGF
3Jnz8JEduTO3ueQ8AOVSueRnP+YXTNP5KC3Tv4jNDwAAAP//AwBQSwMEFAAGAAgAAAAhAMFeIUVnAQAAnQIAABEACAFkb2NQcm9w
cy9jb3JlLnhtbCCiBAEooAABAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAHySSW+DMBSE75X6H5DvxCYLjSxC1EU5NVLU
UnW5ufZLYhWMZTtN0l9fA4HSRT3CzHyaeXIyPxR58A7GylLNUDQgKADFSyHVZoYeskU4RYF1TAmWlwpm6AgWzdPzs4RryksDK1Nq
ME6CDTxJWcr1DG2d0xRjy7dQMDvwDuXFdWkK5vyn2WDN+BvbAB4SEuMCHBPMMVwBQ90R0QkpeIfUO5PXAMEx5FCAchZHgwh/eR2Y
wv4ZqJWes5DuqP2mU90+W/BG7NwHKzvjfr8f7Ed1Dd8/wk/L2/t6aihVdSsOKE0Ep9wAc6VJ79iHXAfPO7WzCe79r26YM+uW/txr
CeLq6K3HLVPB5athJsG/9TayMlI5EOmQDOOQxOGIZCSi45iOxi9drjX5LvX0phCIwI+hzfRWeRxd32QLVPHGIZmEUZyRKSUxjWLP
+5GvxjXA4tT8f2K/4QUlkx6xBaR16e8PKv0EAAD//wMAUEsDBBQABgAIAAAAIQBtUYYHiwEAABIDAAAQAAgBZG9jUHJvcHMvYXBw
LnhtbCCiBAEooAABAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAJySQW/bMAyF7wP2HwzdE9lpUQyBrGJIO/SwYQGStmdN
pmOhsmSIrJHs14+20dRZd9qN5Ht4+kRJ3R5bn/WQ0MVQimKZiwyCjZULh1I87r8tvogMyYTK+BigFCdAcas/f1LbFDtI5AAzjghY
ioaoW0uJtoHW4JLlwEodU2uI23SQsa6dhbtoX1sIJFd5fiPhSBAqqBbdOVBMieue/je0inbgw6f9qWNgrb52nXfWEN9S/3A2RYw1
ZfdHC17JuaiYbgf2NTk66VzJeat21njYcLCujUdQ8n2gHsAMS9sal1CrntY9WIopQ/eb17YS2S+DMOCUojfJmUCMNdimZqx9h5T0
c0wv2AAQKsmGaTiWc++8dte6GA1cXBqHgAmEhUvEvSMP+LPemkT/IC7mxCPDxDvh5MVila+uPwCOd+aj/grfxLYz4cTCufruwgs+
dvt4Zwje9nk5VLvGJKj4Cc77Pg/UA68y+SFk05hwgOrN81EYXv9p+uK6uFnmVzk/7Gym5Ptn1n8AAAD//wMAUEsBAi0AFAAGAAgA
AAAhAHQ2WqZ6AQAAhAUAABMAAAAAAAAAAAAAAAAAAAAAAFtDb250ZW50X1R5cGVzXS54bWxQSwECLQAUAAYACAAAACEAtVUwI/QA
AABMAgAACwAAAAAAAAAAAAAAAACzAwAAX3JlbHMvLnJlbHNQSwECLQAUAAYACAAAACEAQIl22KUDAAAcCQAADwAAAAAAAAAAAAAA
AADYBgAAeGwvd29ya2Jvb2sueG1sUEsBAi0AFAAGAAgAAAAhAJIHlOwEAQAAPwMAABoAAAAAAAAAAAAAAAAAqgoAAHhsL19yZWxz
L3dvcmtib29rLnhtbC5yZWxzUEsBAi0AFAAGAAgAAAAhAPAXgxC/DAAANkUAABgAAAAAAAAAAAAAAAAA7gwAAHhsL3dvcmtzaGVl
dHMvc2hlZXQxLnhtbFBLAQItABQABgAIAAAAIQDBFxC+TgcAAMYgAAATAAAAAAAAAAAAAAAAAOMZAAB4bC90aGVtZS90aGVtZTEu
eG1sUEsBAi0AFAAGAAgAAAAhAPYUe8faBAAALhIAAA0AAAAAAAAAAAAAAAAAYiEAAHhsL3N0eWxlcy54bWxQSwECLQAUAAYACAAA
ACEAxikDNJ0CAACrBgAAFAAAAAAAAAAAAAAAAABnJgAAeGwvc2hhcmVkU3RyaW5ncy54bWxQSwECLQAUAAYACAAAACEAO20yS8EA
AABCAQAAIwAAAAAAAAAAAAAAAAA2KQAAeGwvd29ya3NoZWV0cy9fcmVscy9zaGVldDEueG1sLnJlbHNQSwECLQAUAAYACAAAACEA
nifDJrsBAAA0FQAAJwAAAAAAAAAAAAAAAAA4KgAAeGwvcHJpbnRlclNldHRpbmdzL3ByaW50ZXJTZXR0aW5nczEuYmluUEsBAi0A
FAAGAAgAAAAhAIvuree/AQAAiggAABAAAAAAAAAAAAAAAAAAOCwAAHhsL2NhbGNDaGFpbi54bWxQSwECLQAUAAYACAAAACEAwV4h
RWcBAACdAgAAEQAAAAAAAAAAAAAAAAAlLgAAZG9jUHJvcHMvY29yZS54bWxQSwECLQAUAAYACAAAACEAbVGGB4sBAAASAwAAEAAA
AAAAAAAAAAAAAADDMAAAZG9jUHJvcHMvYXBwLnhtbFBLBQYAAAAADQANAGQDAACEMwAAAAA=
"""
