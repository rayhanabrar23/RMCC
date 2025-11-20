import streamlit as st
import pandas as pd
from datetime import date
import io
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, numbers, Border, Side

# --- Konfigurasi File Lokal ---
# Pastikan file ini ada di root folder dengan header yang benar
BORROW_FILE = 'borrow_contracts.csv'
RETURN_FILE = 'return_events.csv'

# --- Fungsi Utility Data Loading (CSV) ---

@st.cache_data(ttl=600) 
def load_data(file_path):
    """Memuat data dari file CSV lokal atau membuat DataFrame kosong."""
    try:
        df = pd.read_csv(file_path, encoding='utf-8') 
        
        # Konversi tipe data tanggal
        date_cols = ['Request Date', 'Reimbursement Date', 'Actual Return Date', 'Original Request Date']
        for col in date_cols:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors='coerce').dt.date
                
        # Konversi numerik
        for col in ['Borrow Amount (shares)', 'Borrow Price', 'Return Shares']:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col].astype(str).str.replace(',', '.'), errors='coerce').fillna(0)

        return df.dropna(how='all')
    except (FileNotFoundError, pd.errors.EmptyDataError):
        if file_path == BORROW_FILE:
            cols = ['Request Date', 'Borrower', 'Stock Code', 'Borrow Amount (shares)', 'Borrow Price', 'Reimbursement Date', 'Status']
        elif file_path == RETURN_FILE:
            cols = ['Original Request Date', 'Borrower', 'Stock Code', 'Return Shares', 'Actual Return Date']
        else:
            cols = []
            
        return pd.DataFrame(columns=cols)
    except Exception as e:
        st.error(f"Gagal memuat data dari CSV. Pastikan header sudah benar. Error: {e}")
        return pd.DataFrame()

# --- Fungsi Utility Data Saving (CSV) ---

def append_and_save(file_path, new_data):
    """Menambahkan data baru, menyimpan kembali ke CSV, dan memperbarui cache."""
    
    df = load_data(file_path)
    new_df = pd.concat([df, pd.DataFrame([new_data])], ignore_index=True)
    
    try:
        new_df.to_csv(file_path, index=False, encoding='utf-8')
    except Exception as e:
        st.error(f"Gagal menyimpan ke file CSV lokal. Error: {e}")
        return False

    load_data.clear() 
    return True

# --- FUNGSI MENGOSONGKAN DATABASE ---
def clear_local_csv(file_path):
    """Mengosongkan file CSV lokal dengan menulis DataFrame kosong ke dalamnya."""
    try:
        if file_path == BORROW_FILE:
            cols = ['Request Date', 'Borrower', 'Stock Code', 'Borrow Amount (shares)', 'Borrow Price', 'Reimbursement Date', 'Status']
        elif file_path == RETURN_FILE:
            cols = ['Original Request Date', 'Borrower', 'Stock Code', 'Return Shares', 'Actual Return Date']
        else:
            cols = []
            
        empty_df = pd.DataFrame(columns=cols)
        empty_df.to_csv(file_path, index=False, encoding='utf-8')
        return True
    except Exception as e:
        if not os.path.exists(file_path):
            return True 
        st.error(f"Gagal mengosongkan file CSV: {file_path}. Error: {e}")
        return False
# --- AKHIR FUNGSI MENGOSONGKAN DATABASE ---


# --- Fungsi: Membaca Data Lent dari Template Excel yang Diunggah ---

def read_lent_data_from_template(uploaded_file):
    """Membaca data Lent yang sudah ada di template (Baris 8 hingga 18)."""
    try:
        uploaded_file.seek(0)
        wb = load_workbook(uploaded_file, data_only=True) 
        ws = wb.active
        
        data = []
        for r in range(8, 19): 
            row_data = [ws.cell(row=r, column=c).value for c in range(1, 10)]
            
            if row_data[0] is None or (isinstance(row_data[0], str) and 'Total' in row_data[0]):
                break
            
            data.append(row_data)

        cols = ['Request Date', 'Borrower', 'Stock Code', 'Borrow Amount (shares)', 
                'Borrow Price', 'Borrow Value', 'Status', 'Reimbursement Date', 
                'Estimated Period (days)']
        
        df = pd.DataFrame(data, columns=cols)
        
        if not df.empty and isinstance(df.iloc[0]['Request Date'], str):
             df = df.iloc[1:].copy()

        for col in ['Request Date', 'Reimbursement Date']:
             df[col] = pd.to_datetime(df[col], errors='coerce').dt.date
        for col in ['Borrow Amount (shares)', 'Borrow Price', 'Borrow Value']:
             df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
             
        return df.dropna(how='all')

    except Exception as e:
        return pd.DataFrame() 
        
# --- Fungsi Logika Pembuatan Laporan ---

def generate_report_dfs(borrow_df, return_df, report_date):
    """Memproses data pinjaman dan pengembalian untuk menghasilkan 2 tabel laporan."""
    
    # 1. Proses Data Pinjaman (LENT) - Tabel Atas
    lent_df = borrow_df.copy()
    lent_df['Borrow Value'] = lent_df['Borrow Amount (shares)'] * lent_df['Borrow Price']
    
    lent_df['Request Date'] = pd.to_datetime(lent_df['Request Date'], errors='coerce')
    lent_df['Reimbursement Date'] = pd.to_datetime(lent_df['Reimbursement Date'], errors='coerce')

    lent_df['Estimated Period (days)'] = (lent_df['Reimbursement Date'] - lent_df['Request Date']).dt.days
    lent_df['Status'] = 'Lent'
    
    lent_df = lent_df[['Request Date', 'Borrower', 'Stock Code', 'Borrow Amount (shares)', 'Borrow Price', 'Borrow Value', 'Status', 'Reimbursement Date', 'Estimated Period (days)']]
    
    lent_df['Request Date'] = lent_df['Request Date'].dt.date
    lent_df['Reimbursement Date'] = lent_df['Reimbursement Date'].dt.date
    
    # 2. Proses Data Pengembalian (RETURNED) - Tabel Bawah
    
    returned_on_date = return_df.copy() 
    
    if returned_on_date.empty:
        return lent_df, pd.DataFrame()

    returned_on_date['Original Request Date'] = pd.to_datetime(returned_on_date['Original Request Date'], errors='coerce')
    returned_on_date['Actual Return Date'] = pd.to_datetime(returned_on_date['Actual Return Date'], errors='coerce')

    returned_on_date = returned_on_date[returned_on_date['Actual Return Date'].dt.date <= report_date].copy()

    merge_cols = ['Stock Code', 'Borrower']
    merged_returns = returned_on_date.merge(
        borrow_df[['Stock Code', 'Borrower', 'Borrow Price']].drop_duplicates(subset=merge_cols, keep='last'),
        on=merge_cols, how='left'
    )
    
    merged_returns['Estimated Period (days)'] = (
        merged_returns['Actual Return Date'] - 
        merged_returns['Original Request Date']
    ).dt.days

    returned_df_final = pd.DataFrame({
        'Request Date': merged_returns['Original Request Date'].dt.date,
        'Borrower': merged_returns['Borrower'],
        'Stock Code': merged_returns['Stock Code'],
        'Borrow Amount (shares)': merged_returns['Return Shares'], 
        'Borrow Price': merged_returns['Borrow Price'].fillna(0),
        'Status': 'Returned',
        'Reimbursement Date': merged_returns['Actual Return Date'].dt.date,
        'Estimated Period (days)': merged_returns['Estimated Period (days)']
    })

    returned_df_final['Borrow Value'] = returned_df_final['Borrow Amount (shares)'] * returned_df_final['Borrow Price']
    
    return lent_df, returned_df_final.reset_index(drop=True)

# --- Fungsi Pembuatan Excel (OpenPyXL, REVISI AKHIR STYLING) ---

def create_excel_report(template_file, lent_df, returned_df, report_date):
    """Mengisi template Excel dengan strategi append, mempertahankan format, dan menerapkan style baru."""
    
    try:
        template_file.seek(0)
        wb = load_workbook(template_file)
        ws = wb.active 
    except Exception as e:
        st.error(f"Gagal memuat template Excel: {e}")
        return None

    # Style baru untuk data
    NEW_DATA_FONT = Font(name='Roboto Condensed', size=9) 
    THIN_BORDER = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center') 

    # --- KONFIGURASI BARIS UTAMA ---
    START_ROW_LENT_DATA = 8             
    ORIGINAL_TOTAL_LENT_ROW = 19        
    ROWS_BETWEEN_TOTAL_LENT_AND_RETURN_HEADER = 6 
    
    # --- 1. PEMBARUAN HEADER TANGGAL (ROW 2) ---
    date_str_formatted = report_date.strftime("%d-%b-%Y")
    
    new_header_value = f"Daily As of Date: {date_str_formatted} – {date_str_formatted}"
    
    ws['A2'].value = new_header_value
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # --- 2. PERGESERAN BARIS LENT ---
    data_in_df_lent = len(lent_df) 
    old_data_count = ORIGINAL_TOTAL_LENT_ROW - START_ROW_LENT_DATA
    rows_to_write = data_in_df_lent + 1 

    row_difference = rows_to_write - old_data_count
    
    if row_difference > 0:
        ws.insert_rows(ORIGINAL_TOTAL_LENT_ROW, row_difference)
    elif row_difference < 0:
        ws.delete_rows(ORIGINAL_TOTAL_LENT_ROW + row_difference, abs(row_difference))


    # --- 3. TULIS ULANG SELURUH DATA LENT (Mulai dari Baris 9) ---
    
    data_to_write = dataframe_to_rows(lent_df, header=False, index=False)
    
    for r_idx, row in enumerate(data_to_write):
        row_num = START_ROW_LENT_DATA + 1 + r_idx 
        
        for c_idx in range(1, 10):
            if c_idx - 1 < len(row):
                 cell = ws.cell(row=row_num, column=c_idx, value=row[c_idx - 1])
            
            # Terapkan Font dan Border
            cell.font = NEW_DATA_FONT
            cell.border = THIN_BORDER

            # Set default alignment to Center 
            cell.alignment = CENTER_ALIGNMENT 
            
            # Format Angka (Kolom D, E, F)
            if c_idx in [4, 5, 6]:
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                # Override: Numerical data should be right-aligned
                cell.alignment = Alignment(horizontal='right', vertical='center')
            
            # Format Tanggal (Kolom A, H)
            if c_idx in [1, 8]:
                cell.number_format = 'DD-MMM-YY'


    # --- 4. PERBARUI BARIS TOTAL LENT ---
    
    new_total_lent_row = START_ROW_LENT_DATA + data_in_df_lent + 1
    
    for col_idx in range(1, 10):
        source_cell = ws.cell(row=ORIGINAL_TOTAL_LENT_ROW, column=col_idx)
        dest_cell = ws.cell(row=new_total_lent_row, column=col_idx)
        dest_cell._style = source_cell._style 
        
        if col_idx == 6: 
            dest_cell.value = f'=SUM(F{START_ROW_LENT_DATA + 1}:F{new_total_lent_row - 1})'
        elif col_idx == 1: 
             dest_cell.value = 'Total'
        else:
            dest_cell.value = None


    # --- 5. TULIS DATA RETURNED (TABEL BAWAH) ---
    
    NEW_START_ROW_RETURNED_HEADER = new_total_lent_row + ROWS_BETWEEN_TOTAL_LENT_AND_RETURN_HEADER 
    START_ROW_RETURNED_DATA = NEW_START_ROW_RETURNED_HEADER + 2 

    ws.delete_rows(START_ROW_RETURNED_DATA, 100) 
    
    new_returned_data_count = len(returned_df)
    if new_returned_data_count > 0:
        ws.insert_rows(START_ROW_RETURNED_DATA, new_returned_data_count)
        
    # Tulis data RETURNED baru
    for r_idx, row in enumerate(dataframe_to_rows(returned_df, header=False, index=False)):
        row_num = START_ROW_RETURNED_DATA + r_idx
        
        for c_idx in range(1, 10):
             cell = ws.cell(row=row_num, column=c_idx, value=row[c_idx - 1])
             
             # Terapkan Font dan Border
             cell.font = NEW_DATA_FONT
             cell.border = THIN_BORDER

             # Set default alignment to Center
             cell.alignment = CENTER_ALIGNMENT
             
             # Format Angka (Kolom D, E, F)
             if c_idx in [4, 5, 6]:
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                # Override: Numerical data should be right-aligned
                cell.alignment = Alignment(horizontal='right', vertical='center')
             
             # Format Tanggal (Kolom A, H)
             if c_idx in [1, 8]:
                cell.number_format = 'DD-MMM-YY'
        
    # --- 6. PERBARUI BARIS TOTAL RETURNED ---
    
    new_total_returned_row = START_ROW_RETURNED_DATA + new_returned_data_count 
    
    ws.cell(row=new_total_returned_row, column=6).value = f'=SUM(F{START_ROW_RETURNED_DATA}:F{new_total_returned_row - 1})'
    ws.cell(row=new_total_returned_row, column=6).font = Font(bold=True)
    ws.cell(row=new_total_returned_row, column=6).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
    
    ws.cell(row=new_total_returned_row, column=1).value = 'Total'
    ws.cell(row=new_total_returned_row, column=1).font = Font(bold=True)
    
    ws.cell(row=new_total_returned_row, column=5).value = None
    ws.cell(row=new_total_returned_row, column=7).value = None


    # --- SIMPAN DAN KEMBALIKAN FILE ---
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# --- Aplikasi Streamlit Utama ---

st.set_page_config(layout="wide", page_title="SLB Daily Position Generator")

st.title("📊 SLB Daily Position Automation (Gabungan DB)")

# --- Pemuatan Data dari CSV ---
if 'slb_borrow_df' not in st.session_state:
    st.session_state['slb_borrow_df'] = load_data(BORROW_FILE)
if 'slb_return_df' not in st.session_state:
    st.session_state['slb_return_df'] = load_data(RETURN_FILE)

borrow_df = st.session_state['slb_borrow_df']
return_df = st.session_state['slb_return_df']


# --- Antarmuka Pengguna ---

uploaded_file = st.file_uploader("1. Upload Template Excel SLB Anda", type=['xlsx'])
report_date = st.date_input("2. Tanggal Laporan Posisi", date.today())
st.markdown("---")

col1, col2 = st.columns(2)

# --- Form Input Peminjaman Baru (LENT) ---
with col1:
    st.header("➕ Input Kontrak Pinjaman Baru (LENT)")
    with st.form("borrow_form", clear_on_submit=True):
        req_date = st.date_input("Request Date", date.today())
        borrower = st.text_input("Borrower (cth: HD, KPEI)")
        stock = st.text_input("Stock Code (cth: PGEO, MEDC)")
        amount = st.number_input("Borrow Amount (shares)", min_value=1, step=100)
        price = st.number_input("Borrow Price", min_value=0.01, step=0.001, format="%.3f")
        reimburse_date = st.date_input("Reimbursement Date (Jatuh Tempo)", date.today())
        
        submitted = st.form_submit_button("Simpan Kontrak Pinjaman ke CSV Lokal")
        if submitted:
            if all([borrower, stock, amount, price]):
                new_borrow = {
                    'Request Date': req_date,
                    'Borrower': borrower.upper(),
                    'Stock Code': stock.upper(),
                    'Borrow Amount (shares)': amount,
                    'Borrow Price': price,
                    'Reimbursement Date': reimburse_date,
                    'Status': 'Lent'
                }
                if append_and_save(BORROW_FILE, new_borrow):
                    st.session_state['slb_borrow_df'] = load_data(BORROW_FILE)
                    st.success("Kontrak Pinjaman Baru berhasil ditambahkan ke CSV lokal!")
                
            else:
                st.error("Mohon isi semua kolom input Pinjaman.")

# --- Form Input Pengembalian (RETURNED) ---
with col2:
    st.header("➖ Input Event Pengembalian (RETURNED)")
    with st.form("return_form", clear_on_submit=True):
        actual_return_date = st.date_input("Actual Return Date (Tanggal Keluar)", date.today())
        
        active_contracts = borrow_df.copy()
        if not active_contracts.empty:
            active_contracts['Contract Key'] = active_contracts['Stock Code'] + ' | ' + active_contracts['Borrower']
            stock_options = ['Pilih Saham | Peminjam'] + active_contracts['Contract Key'].unique().tolist()
        else:
            stock_options = ['Tidak ada kontrak aktif']

        selected_contract = st.selectbox("Pilih Kontrak Saham (Kode Saham | Peminjam)", stock_options)
        
        return_shares = st.number_input("Return Amount (shares)", min_value=1, step=100)
        
        return_submitted = st.form_submit_button("Simpan Event Pengembalian ke CSV Lokal")
        
        if return_submitted:
            if selected_contract != 'Pilih Saham | Peminjam' and selected_contract != 'Tidak ada kontrak aktif' and return_shares:
                try:
                    stock_code, borrower_name = selected_contract.split(' | ')
                except ValueError:
                    st.error("Format kontrak tidak valid.")
                    st.stop()
                
                original_contract = borrow_df[
                    (borrow_df['Stock Code'] == stock_code) & 
                    (borrow_df['Borrower'] == borrower_name)
                ]
                
                if original_contract.empty:
                    st.error("Kontrak asli tidak ditemukan.")
                    st.stop()
                    
                original_contract = original_contract.sort_values(by='Reimbursement Date', ascending=False).iloc[0]

                new_return = {
                    'Original Request Date': original_contract['Request Date'],
                    'Borrower': borrower_name,
                    'Stock Code': stock_code,
                    'Return Shares': return_shares,
                    'Actual Return Date': actual_return_date
                }
                
                if append_and_save(RETURN_FILE, new_return):
                    st.session_state['slb_return_df'] = load_data(RETURN_FILE)
                    st.success(f"Event Pengembalian untuk {stock_code} berhasil ditambahkan ke CSV lokal!")
            else:
                st.error("Mohon pilih kontrak dan isi jumlah saham.")

st.markdown("---")

# --- Bagian Pembuatan Laporan ---

if uploaded_file is not None:
    st.header("3. Hasil Laporan dan Download")
    
    existing_lent_df = read_lent_data_from_template(uploaded_file)
    csv_lent_df = st.session_state['slb_borrow_df']
    
    merge_cols = ['Request Date', 'Borrower', 'Stock Code', 'Borrow Amount (shares)', 'Borrow Price', 'Reimbursement Date']

    existing_lent_df = existing_lent_df[~existing_lent_df.set_index(merge_cols).index.isin(csv_lent_df.set_index(merge_cols).index)]

    combined_lent_df = pd.concat([existing_lent_df, csv_lent_df], ignore_index=True)

    
    final_lent_df, final_returned_df = generate_report_dfs(combined_lent_df, st.session_state['slb_return_df'], report_date)
    
    uploaded_file.seek(0)
    
    excel_report = create_excel_report(uploaded_file, final_lent_df, final_returned_df, report_date)

    if excel_report:
        file_name = f"SLB Position {report_date.strftime('%Y%m%d')}.xlsx"
        
        # Tombol Download ditempatkan di sini
        st.download_button(
            label="Download Laporan Excel Otomatis",
            data=excel_report,
            file_name=file_name,
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # --- BLOK REVISI: MENGOSONGKAN DATABASE SECARA OTOMATIS ---
        # Gunakan query param atau teknik lain untuk pemicu, tapi yang paling sederhana adalah
        # memberikan notifikasi bahwa user harus me-refresh setelah download.
        
        st.markdown("---")
        if not csv_lent_df.empty or not return_df.empty:
            st.warning("**PENTING:** Data input baru (Lent/Return) telah dimasukkan ke dalam laporan ini. Untuk MENGHINDARI DUPLIKASI di laporan berikutnya, **mohon tekan tombol di bawah**.")
            
            if st.button("4. Kosongkan Database Input Lokal Sekarang"):
                if clear_local_csv(BORROW_FILE) and clear_local_csv(RETURN_FILE):
                    # Perbarui session state agar input form berikutnya kosong
                    st.session_state['slb_borrow_df'] = load_data(BORROW_FILE)
                    st.session_state['slb_return_df'] = load_data(RETURN_FILE)
                    st.success("✅ Database Pinjaman dan Pengembalian lokal **berhasil dikosongkan**! Silakan **refresh halaman** untuk mulai input data baru.")
                    st.balloons()
                else:
                    st.error("❌ Gagal membersihkan database lokal.")
        else:
             st.info("Database input lokal sudah kosong.")
        # --- AKHIR BLOK REVISI ---
        
    else:
        st.warning("Data belum tersedia atau gagal memproses laporan.")
