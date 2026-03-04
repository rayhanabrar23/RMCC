import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
import io

st.set_page_config(page_title="Finance Report Linkage", layout="wide")

st.title("📊 Disburse & Repayment Automator")
st.markdown("Unggah file master dan file broker (XLS/XML), lalu sistem akan memproses otomatis.")

# --- KONFIGURASI DINAMIS ---
with st.sidebar:
    st.header("Konfigurasi")
    bulan_dipilih = st.selectbox("Pilih Bulan Proses", 
                                ["JANUARI", "FEBRUARI", "MARET", "APRIL", "MEI", "JUNI", 
                                 "JULI", "AGUSTUS", "SEPTEMBER", "OKTOBER", "NOVEMBER", "DESEMBER"])
    tahun_proses = st.text_input("Tahun", "2026")

def get_column_letter(month_name):
    mapping = {
        'JANUARI': 'F', 'FEBRUARI': 'G', 'MARET': 'H', 'APRIL': 'I',
        'MEI': 'J', 'JUNI': 'K', 'JULI': 'L', 'AGUSTUS': 'M',
        'SEPTEMBER': 'N', 'OKTOBER': 'O', 'NOVEMBER': 'P', 'DESEMBER': 'Q'
    }
    return mapping.get(month_name.upper(), 'F')

def smart_read_excel(uploaded_file):
    """Fungsi untuk membaca file Excel standar atau XML Spreadsheet 2003"""
    content = uploaded_file.read()
    try:
        # Coba baca sebagai Excel standar (xlsx/xls)
        return pd.read_excel(BytesIO(content), header=None)
    except Exception:
        try:
            # Jika gagal, coba baca sebagai XML (sering terjadi pada file .xls hasil export sistem)
            return pd.read_html(BytesIO(content))[0]
        except Exception as e:
            st.error(f"Gagal membaca file {uploaded_file.name}: Format tidak didukung.")
            return None

# --- UPLOAD FILE ---
col1, col2 = st.columns(2)

with col1:
    master_file = st.file_uploader("1. Unggah File Master (Template .xlsx)", type=['xlsx'])

with col2:
    broker_files = st.file_uploader("2. Unggah Semua File Broker (.xls / .xlsx)", type=['xls', 'xlsx'], accept_multiple_files=True)

if master_file and broker_files:
    if st.button("🚀 Proses & Generate Laporan"):
        try:
            # Load Master Workbook
            wb = openpyxl.load_workbook(master_file)
            ws_main = wb['Disburse & Repay Jan-Des']
            
            target_col = get_column_letter(bulan_dipilih)
            row_map_disburse = {'EP': 5, 'HD': 6, 'HP': 7, 'XC': 8}
            row_map_repayment = {'EP': 16, 'HD': 17, 'HP': 18, 'XC': 19}
            
            process_log = []

            for f in broker_files:
                fname = f.name.upper()
                
                # Deteksi Broker
                code = None
                for c in ['EP', 'HD', 'HP', 'XC']:
                    if f" {c}" in fname or f"{c}." in fname or fname.startswith(c):
                        code = c
                        break
                
                if not code:
                    process_log.append({"File": f.name, "Status": "⚠️ Skip", "Keterangan": "Broker tidak terdeteksi"})
                    continue

                # Gunakan fungsi smart_read
                df_raw = smart_read_excel(f)
                
                if df_raw is None:
                    continue

                val = 0
                if "DISBURSE" in fname:
                    # Cari 'Grand Total' di seluruh dataframe
                    mask_total = df_raw.astype(str).apply(lambda x: x.str.contains('Grand Total', case=False)).any(axis=1)
                    df_total_row = df_raw[mask_total]
                    if not df_total_row.empty:
                        numeric_only = pd.to_numeric(df_total_row.iloc[-1], errors='coerce').dropna()
                        val = numeric_only.iloc[-1] if not numeric_only.empty else 0
                    
                    ws_main[f"{target_col}{row_map_disburse[code]}"] = val
                    process_log.append({"File": f.name, "Status": "✅ Berhasil", "Keterangan": f"Disburse {code}: {val:,.0f}"})

                elif "REPAYMENT" in fname:
                    # Ambil kolom ke-9 (Index 8). Jika XML, mungkin perlu penyesuaian index
                    try:
                        col_target = df_raw.iloc[:, 8] # Kolom ke-9
                        col_numeric = pd.to_numeric(col_target, errors='coerce').dropna()
                        val = col_numeric.iloc[-1] if not col_numeric.empty else 0
                    except:
                        val = 0
                    
                    ws_main[f"{target_col}{row_map_repayment[code]}"] = val
                    process_log.append({"File": f.name, "Status": "✅ Berhasil", "Keterangan": f"Repayment {code}: {val:,.0f}"})

            # Update Summary Tahunan
            s_sum_name = "Summary Tahunan"
            if s_sum_name in wb.sheetnames:
                ws_sum = wb[s_sum_name]
                month_list = ['JANUARI', 'FEBRUARI', 'MARET', 'APRIL', 'MEI', 'JUNI', 
                              'JULI', 'AGUSTUS', 'SEPTEMBER', 'OKTOBER', 'NOVEMBER', 'DESEMBER']
                summary_row = month_list.index(bulan_dipilih.upper()) + 2
                ws_sum[f'B{summary_row}'] = f"=SUM('Disburse & Repay Jan-Des'!{target_col}5:{target_col}8)"

            # Simpan ke Memory
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            st.success(f"Pemrosesan Selesai!")
            st.table(pd.DataFrame(process_log))

            st.download_button(
                label="📥 Download File Master Terbaru",
                data=output,
                file_name=f"Update_{bulan_dipilih}_{tahun_proses}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"Terjadi kesalahan teknis: {e}")
