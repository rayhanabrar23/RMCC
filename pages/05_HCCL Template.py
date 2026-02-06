import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook

st.set_page_config(page_title="Margin Merger - Ultra Low RAM", layout="centered")

st.title("📊 Margin Trades Merger (Low Memory Mode)")
st.info("Mode ini mencicil data satu per satu agar tidak crash.")

template_file = st.file_uploader("1. Upload Template (.xlsx)", type=['xlsx'])
daily_files = st.file_uploader("2. Upload CSV Harian", type=['csv'], accept_multiple_files=True)

if st.button("🚀 Proses Secara Bertahap"):
    if template_file and daily_files:
        try:
            with st.spinner('Sedang menulis data ke Excel (ini mungkin butuh waktu)...'):
                # 1. Load Template ke Memory
                template_bytes = template_file.read()
                book = load_workbook(BytesIO(template_bytes))
                
                # Pastikan sheet Januari ada, atau buat jika tidak ada
                if 'Januari' not in book.sheetnames:
                    book.create_sheet('Januari')
                
                sheet = book['Januari']
                
                # Cek jika sheet kosong, kita akan tulis header nanti
                # Jika sudah ada isinya, kita cari baris terakhir
                start_row = sheet.max_row
                
                sorted_files = sorted(daily_files, key=lambda x: x.name)
                progress_bar = st.progress(0)

                for i, f in enumerate(sorted_files):
                    # Baca satu CSV saja
                    df = pd.read_csv(f, sep=None, engine='python')
                    
                    # Jika ini file pertama dan sheet Januari masih kosong, tulis header
                    if i == 0 and start_row <= 1:
                        columns = df.columns.tolist()
                        for col_num, column_title in enumerate(columns, 1):
                            sheet.cell(row=1, column=col_num).value = column_title
                        start_row = 1

                    # Tulis data CSV ke baris berikutnya di Excel
                    # Kita konversi dataframe ke list of lists agar cepat
                    for r_idx, row in enumerate(df.values.tolist(), start_row + 1):
                        for c_idx, value in enumerate(row, 1):
                            sheet.cell(row=r_idx, column=c_idx).value = value
                    
                    # Update baris awal untuk file berikutnya
                    start_row += len(df)
                    
                    # Update Progress & Hapus df dari RAM
                    progress_bar.progress((i + 1) / len(sorted_files))
                    del df 

                # 2. Simpan hasil akhir
                output = BytesIO()
                book.save(output)
                processed_data = output.getvalue()

            st.success(f"Selesai! Berhasil menggabungkan {len(daily_files)} file.")
            
            st.download_button(
                label="📥 Download Hasil",
                data=processed_data,
                file_name="Updated_2026_Q1_Margin_Trades.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"Error: {e}")
    else:
        st.warning("Upload template dan CSV dulu.")
