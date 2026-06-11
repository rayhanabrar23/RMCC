import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime

# Cek login
if "login_status" not in st.session_state or not st.session_state["login_status"]:
    st.error("🚨 Akses Ditolak! Silakan login di halaman utama terlebih dahulu.")
    st.stop()

# ─────────────────────────────────────────
# CUSTOM CSS
# ─────────────────────────────────────────
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

    .stApp { background-color: #0f1117; color: #e8eaf0; }

    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #1a1d2e 0%, #141624 100%);
        border-right: 1px solid #2a2d3e;
    }

    .card {
        background: #1a1d2e;
        border: 1px solid #2a2d3e;
        border-radius: 12px;
        padding: 20px 24px;
        margin-bottom: 16px;
    }
    .card-title {
        font-size: 0.75rem;
        font-weight: 600;
        letter-spacing: 0.08em;
        color: #5c7cfa;
        text-transform: uppercase;
        margin-bottom: 6px;
    }

    .main-header {
        background: linear-gradient(135deg, #1a1d2e 0%, #1e2340 100%);
        border: 1px solid #2a2d3e;
        border-radius: 14px;
        padding: 24px 28px;
        margin-bottom: 24px;
    }
    .main-header h1 { font-size: 1.6rem; font-weight: 700; color: #e8eaf0; margin: 0; }
    .main-header p  { color: #6b7080; margin: 6px 0 0 0; font-size: 0.88rem; }

    .stButton > button {
        background: linear-gradient(135deg, #3b5bdb, #5c7cfa);
        color: white;
        border: none;
        border-radius: 8px;
        font-weight: 600;
        font-size: 0.85rem;
        padding: 10px 24px;
        transition: opacity 0.2s;
    }
    .stButton > button:hover { opacity: 0.88; }

    [data-testid="stFileUploader"] {
        background: #1a1d2e;
        border: 1px dashed #3b5bdb;
        border-radius: 10px;
        padding: 8px;
    }

    hr { border-color: #2a2d3e; }

    .stTextInput > div > div, .stSelectbox > div > div {
        background: #1a1d2e !important;
        border-color: #2a2d3e !important;
        color: #e8eaf0 !important;
    }
</style>
""", unsafe_allow_html=True)

# ============================
# KONSTANTA & KONFIGURASI
# ============================
REPO_KEY_COL = 'Instrument Code'
PHEI_KEY_COL = 'SERIES'
PHEI_VALUE_COL = 'TODAY FAIR PRICE'
HEADER_ROW_INDEX = 9
START_ROW_EXCEL = 11

# ============================
# FUNGSI PEMBERSIH KUNCI
# ============================
def clean_key_extreme(series):
    return series.astype(str).str.strip().str.upper().replace(r'[^A-Z0-9]', '', regex=True)

# ============================
# FUNGSI PENGOLAHAN DATA UTAMA
# ============================
def process_repo_data(df_repo_main, df_phei_lookup):
    st.info("Sedang mencocokkan data Instrument Code dengan Series PHEI...")

    df_repo_main[REPO_KEY_COL] = clean_key_extreme(df_repo_main[REPO_KEY_COL].fillna(''))
    df_phei_lookup[PHEI_KEY_COL] = clean_key_extreme(df_phei_lookup[PHEI_KEY_COL].fillna(''))
    df_phei_lookup = df_phei_lookup.drop_duplicates(subset=[PHEI_KEY_COL])

    df_merged = pd.merge(
        df_repo_main,
        df_phei_lookup[[PHEI_KEY_COL, PHEI_VALUE_COL]],
        left_on=REPO_KEY_COL,
        right_on=PHEI_KEY_COL,
        how='left'
    )

    if PHEI_VALUE_COL in df_merged.columns:
        df_merged['Fair Price PHEI'] = pd.to_numeric(df_merged[PHEI_VALUE_COL], errors='coerce')

    return df_merged

# ============================
# MAIN UI
# ============================
def main():
    # Header
    st.markdown("""
    <div class="main-header">
      <h1>🔄 Otomatisasi Repo Daily Position</h1>
      <p>Mengisi kolom <strong>Fair Price PHEI (J)</strong> dan update tanggal <strong>(A2)</strong> secara otomatis</p>
    </div>
    """, unsafe_allow_html=True)

    # Card: Upload File
    st.markdown('<div class="card"><div class="card-title">Upload File</div>', unsafe_allow_html=True)
    col1, col2 = st.columns(2)
    with col1:
        repo_file_upload = st.file_uploader('1. Template Repo', type=['xlsx'])
    with col2:
        phei_lookup_file = st.file_uploader('2. File PHEI Hari Ini', type=['xlsx', 'csv'])
    st.markdown('</div>', unsafe_allow_html=True)

    if repo_file_upload and phei_lookup_file:
        try:
            repo_bytes = repo_file_upload.getvalue()
            df_repo_raw = pd.read_excel(BytesIO(repo_bytes), header=HEADER_ROW_INDEX)
            df_repo_raw.columns = df_repo_raw.columns.str.replace('\n', ' ').str.strip()
            df_data_only = df_repo_raw[df_repo_raw['No'].notna()].copy()

            if phei_lookup_file.name.endswith('.csv'):
                df_phei = pd.read_csv(phei_lookup_file, encoding='latin1')
            else:
                df_phei = pd.read_excel(phei_lookup_file)
            df_phei.columns = df_phei.columns.str.strip()

            if st.button("▶ Jalankan Proses Update"):
                df_result = process_repo_data(df_data_only, df_phei)

                wb = load_workbook(BytesIO(repo_bytes))
                sheet = wb.active

                try:
                    fair_price_col_idx = df_repo_raw.columns.get_loc('Fair Price PHEI') + 1
                except KeyError:
                    st.error("Kolom 'Fair Price PHEI' tidak ditemukan di template!")
                    return

                today_date = datetime.now().strftime('%d %b %Y')
                date_text = f"Daily As of Date : {today_date} - {today_date}"
                sheet.cell(row=2, column=1, value=date_text)

                for i, val in enumerate(df_result['Fair Price PHEI']):
                    current_row = START_ROW_EXCEL + i
                    final_val = val if pd.notna(val) else None
                    sheet.cell(row=current_row, column=fair_price_col_idx, value=final_val)

                output_buffer = BytesIO()
                wb.save(output_buffer)

                st.success(f"✅ Berhasil memproses {len(df_result)} baris data.")

                # Card: Preview
                st.markdown('<div class="card"><div class="card-title">Preview Hasil (Kolom J)</div>', unsafe_allow_html=True)
                st.dataframe(
                    df_result[['No', REPO_KEY_COL, 'Fair Price PHEI']],
                    use_container_width=True
                )
                st.markdown('</div>', unsafe_allow_html=True)

                # Card: Download
                st.markdown('<div class="card"><div class="card-title">Unduh Hasil</div>', unsafe_allow_html=True)
                st.download_button(
                    label="⬇ Unduh File Update",
                    data=output_buffer.getvalue(),
                    file_name=f"Reverse Repo Bonds Daily Position {datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
                st.markdown('</div>', unsafe_allow_html=True)

        except Exception as e:
            st.error(f"Terjadi kesalahan: {e}")

    else:
        st.markdown("""
        <div class="card">
          <div class="card-title">Panduan</div>
          <p style="color:#8a8fa8;font-size:0.85rem;margin:0">
            Upload kedua file di atas, lalu klik
            <strong style="color:#5c7cfa">▶ Jalankan Proses Update</strong> untuk memproses.
          </p>
        </div>
        """, unsafe_allow_html=True)

main()
