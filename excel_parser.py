"""
excel_parser.py
------------------
Baca file Excel "HC dan CL - KOMITE" dan ubah jadi KomiteSummary
yang dipakai ppt_generator.py.

Sheet yang dipakai:
- HC          -> data haircut per saham, dasar kartu ringkasan & alasan perubahan
- GROUP       -> distribusi Group HC (n saham, value, % setelah haircut)
- Saham Baru  -> daftar saham margin baru per bulan
- REF         -> PEI Equity & Max Financing Ratio

Header di sheet sumber punya baris judul/kosong di atas baris header asli,
dan kolom pertama (A) selalu kosong -- makanya lookup kolom dilakukan
by-name (bukan by-index tetap) supaya tahan kalau kolom digeser.
"""

from dataclasses import dataclass, field

from openpyxl import load_workbook


@dataclass
class KomiteSummary:
    periode_label: str
    total_saham: int
    haircut_naik: int
    haircut_turun: int
    uma_count: int
    saham_baru_count: int
    saham_baru_periode_label: str
    pei_equity: float
    max_fin_ratio: float
    reason_counts: dict
    group_dist: list
    total_collateral_value: float


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def _clean(h):
    """Header sel sering mengandung newline ('HAIRCUT\\nJULI...') -> rapikan jadi satu baris."""
    if h is None:
        return ""
    return " ".join(str(h).split())


def _find_header_row(ws, must_contain, max_scan=10):
    """Cari baris header (baris pertama yang mengandung salah satu kata di must_contain)."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        cleaned = [_clean(c).upper() for c in row]
        if any(any(m in c for c in cleaned) for m in must_contain):
            return i + 1, row  # openpyxl row index 1-based
    raise ValueError(f"Header row tidak ditemukan (cari salah satu dari {must_contain})")


def _header_map(row):
    """dict header_bersih -> index (0-based, sesuai posisi di tuple row)."""
    return {_clean(h): i for i, h in enumerate(row) if _clean(h)}


def _col(hmap, *candidates):
    """Ambil index kolom, coba beberapa kemungkinan nama (header sumber suka berubah spasi/urutan kata)."""
    for cand in candidates:
        if cand in hmap:
            return hmap[cand]
    # fallback: cari partial match
    for cand in candidates:
        for h, i in hmap.items():
            if cand.upper() in h.upper():
                return i
    raise KeyError(f"Kolom tidak ditemukan, dicoba: {candidates}")


# ─────────────────────────────────────────────
# PARSER PER SHEET
# ─────────────────────────────────────────────
def _parse_hc_sheet(wb):
    ws = wb["HC"]
    header_row_idx, header_row = _find_header_row(ws, ["KODE EFEK"])
    hmap = _header_map(header_row)

    c_kode = _col(hmap, "KODE EFEK")
    c_uma = _col(hmap, "UMA")
    # kolom diff bulan-ke-bulan namanya berubah tiap periode (mis. "JULI VS JUNI"),
    # jadi dicari yang mengandung kata "VS" DAN "DIFF" tapi BUKAN "KPEI"/"PEI (DIFF)"
    # (dua kolom itu adalah diff PEI-vs-KPEI, bukan diff antar-periode).
    c_diff = None
    for h, i in hmap.items():
        hu = h.upper()
        if "VS" in hu and "DIFF" in hu and "KPEI" not in hu and "PEI (DIFF)" not in hu:
            c_diff = i
            break
    if c_diff is None:
        raise KeyError("Kolom diff haircut antar-periode (mis. 'HAIRCUT JULI VS JUNI (DIFF)') tidak ditemukan")
    c_ket = _col(hmap, "KETERANGAN")
    # ambil kolom KETERANGAN paling kanan (keterangan hasil pembahasan final, bukan bulan lalu)
    ket_cols = [i for h, i in hmap.items() if h.upper() == "KETERANGAN"]
    if ket_cols:
        c_ket = max(ket_cols)

    total_saham = 0
    naik = 0
    turun = 0
    uma_count = 0
    reason_counts = {}

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        kode = row[c_kode]
        if not kode:
            continue
        total_saham += 1

        diff = row[c_diff] or 0
        if diff > 0:
            naik += 1
        elif diff < 0:
            turun += 1

        uma_val = row[c_uma]
        if uma_val not in (None, "-", ""):
            uma_count += 1

        if diff != 0:
            ket = _clean(row[c_ket]) or "Tidak ada keterangan"
            reason_counts[ket] = reason_counts.get(ket, 0) + 1

    return total_saham, naik, turun, uma_count, reason_counts


def _parse_group_sheet(wb):
    ws = wb["GROUP"]
    header_row_idx, header_row = _find_header_row(ws, ["HAIRCUT GROUP"])
    hmap = _header_map(header_row)

    c_name = _col(hmap, "HAIRCUT GROUP")
    c_n = _col(hmap, "NUMBER OF STOCKS IN CLUSTER")
    c_val = _col(hmap, "MAX COLLATERAL VALUE AFTER HAIRCUT")
    c_pct = _col(hmap, "PERCENTAGE AFTER HAIRCUT")

    group_dist = []
    total_value = 0.0
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        # sheet GROUP punya tabel kedua (perbandingan bulan lalu) di bawahnya dengan
        # header yang sama persis -> berhenti begitu ketemu 'TOTAL' di baris manapun,
        # supaya tabel kedua tidak ikut tercampur.
        row_text = [_clean(c).upper() for c in row]
        if "TOTAL" in row_text:
            total_value = row[c_val] if isinstance(row[c_val], (int, float)) else total_value
            break

        name = row[c_name]
        value = row[c_val]
        if not name or not isinstance(value, (int, float)):
            continue
        n_stocks = row[c_n] or 0
        pct = row[c_pct] or 0
        group_dist.append({"name": name, "n_stocks": n_stocks, "value": value, "pct": pct})

    if not total_value:
        total_value = sum(g["value"] for g in group_dist)

    return group_dist, total_value


def _parse_saham_baru_sheet(wb, periode_label):
    ws = wb["Saham Baru"]
    header_row_idx, header_row = _find_header_row(ws, ["BULAN", "KODE EFEK"])
    hmap = _header_map(header_row)
    c_bulan = _col(hmap, "BULAN")

    # nama bulan periode saat ini, mis. "JULI" dari "Juli 2026"
    bulan_key = periode_label.split()[0].upper()

    count = 0
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        bulan = row[c_bulan]
        if bulan and _clean(bulan).upper().startswith(bulan_key):
            count += 1
    return count


def _parse_ref_sheet(wb):
    ws = wb["REF"]
    pei_equity = None
    max_fin_ratio = None
    for row in ws.iter_rows(values_only=True):
        for i, cell in enumerate(row):
            label = _clean(cell).upper()
            if label == "PEI EQUITY" and i + 1 < len(row):
                pei_equity = row[i + 1]
            elif label.startswith("MAX FIN") and i + 1 < len(row):
                max_fin_ratio = row[i + 1]
    if pei_equity is None or max_fin_ratio is None:
        raise ValueError("PEI Equity / Max Fin. Ratio tidak ditemukan di sheet REF")
    return pei_equity, max_fin_ratio


def _detect_periode_label(wb):
    """Cari 'PERIODE: <BULAN> <TAHUN>' di baris judul sheet HC, ubah ke Title Case."""
    ws = wb["HC"]
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        for cell in row:
            if cell and "PERIODE" in _clean(cell).upper():
                label = _clean(cell).split(":")[-1].strip()
                return label.title()
    raise ValueError("Label periode ('PERIODE: ...') tidak ditemukan di sheet HC")


# ─────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────
def parse_komite_excel(file) -> KomiteSummary:
    """file: path atau file-like object (mis. dari st.file_uploader)."""
    wb = load_workbook(file, data_only=True, read_only=True)

    periode_label = _detect_periode_label(wb)
    total_saham, naik, turun, uma_count, reason_counts = _parse_hc_sheet(wb)
    group_dist, total_collateral_value = _parse_group_sheet(wb)
    saham_baru_count = _parse_saham_baru_sheet(wb, periode_label)
    pei_equity, max_fin_ratio = _parse_ref_sheet(wb)

    return KomiteSummary(
        periode_label=periode_label,
        total_saham=total_saham,
        haircut_naik=naik,
        haircut_turun=turun,
        uma_count=uma_count,
        saham_baru_count=saham_baru_count,
        saham_baru_periode_label=periode_label,
        pei_equity=pei_equity,
        max_fin_ratio=max_fin_ratio,
        reason_counts=reason_counts,
        group_dist=group_dist,
        total_collateral_value=total_collateral_value,
    )
