import streamlit as st
# Cek apakah sudah login dari halaman utama
if "login_status" not in st.session_state or not st.session_state["login_status"]:
    st.error("🚨 Akses Ditolak! Silakan login di halaman utama terlebih dahulu.")
    st.stop() # Hentikan aplikasi di sini
from style_utils import apply_custom_style

# Panggil fungsi gaya
apply_custom_style()

# Cek apakah user sudah login (Proteksi Halaman)
if "login_status" not in st.session_state or not st.session_state["login_status"]:
    st.warning("Silakan login terlebih dahulu di halaman utama.")
    st.stop()
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, NamedStyle, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime

# ============================
# KONFIGURASI GLOBAL CL
# ============================
COL_RMCC = 'CONCENTRATION LIMIT USULAN RMCC'
COL_LISTED = 'CONCENTRATION LIMIT TERKENA % LISTED SHARES'
COL_FF = 'CONCENTRATION LIMIT TERKENA % FREE FLOAT'
COL_PERHITUNGAN = 'CONCENTRATION LIMIT SESUAI PERHITUNGAN'
THRESHOLD_5M = 5_000_000_000
# REVISI: KPIG dihapus
KODE_EFEK_KHUSUS = ['LPKR', 'MLPL', 'NOBU', 'PTPP', 'SILO']
# REVISI: KPIG dihapus
OVERRIDE_MAPPING = {
    'LPKR': 10_000_000_000, 'MLPL': 10_000_000_000,
    'NOBU': 10_000_000_000, 'PTPP': 50_000_000_000, 'SILO': 10_000_000_000
}

# Nilai dan toleransi untuk pengecekan 100%
TARGET_100 = 100.0
TOLERANCE = 1e-6 

# ===============================================================
# FUNGSI UTILITAS UNTUK CONCENTRATION LIMIT (CL)
# ===============================================================

def calc_concentration_limit_listed(row):
    try:
        if row['PERBANDINGAN DENGAN LISTED SHARES (Sesuai Perhitungan)'] >= 0.05:
            return 0.0499 * row['LISTED SHARES'] * row['CLOSING PRICE']
        return None
    except Exception:
        return None

def calc_concentration_limit_ff(row):
    try:
        if row['PERBANDINGAN DENGAN FREE FLOAT (Sesuai Perhitungan)'] >= 0.20:
            return 0.1999 * row['FREE FLOAT (DALAM LEMBAR)'] * row['CLOSING PRICE']
        return None
    except Exception:
        return None

# Fungsi override_rmcc_limit (diperbaiki agar CL=0 tetap 0 dan hanya CAP)
def override_rmcc_limit(row):
    kode = row['KODE EFEK']
    nilai_rmcc = row[COL_RMCC]
    
    if kode in OVERRIDE_MAPPING:
        nilai_override = OVERRIDE_MAPPING[kode]
        
        if nilai_rmcc == 0.0:
            return 0.0
            
        return min(nilai_rmcc, nilai_override)
        
    return nilai_rmcc

def reset_concentration_limit(
    df_main: pd.DataFrame,
    haircut_col: str = "HAIRCUT PEI USULAN DIVISI",
    conc_limit_col: str = "CONCENTRATION LIMIT USULAN RMCC",
    conc_calc_col: str = "CONCENTRATION LIMIT SESUAI PERHITUNGAN",
    tolerance: float = 1e-6,
    threshold_limit: float = 5_000_000_000,
    inplace: bool = True
) -> pd.DataFrame:
    if not inplace:
        df_main = df_main.copy()
        
    if haircut_col not in df_main.columns:
        df_main[haircut_col] = 0.0
        
    df_main[haircut_col] = pd.to_numeric(df_main[haircut_col], errors='coerce')
    df_main[conc_calc_col] = pd.to_numeric(df_main[conc_calc_col], errors='coerce')

    valid_haircut = df_main[haircut_col].dropna()
    target_100_reset = 100.0 if not valid_haircut.empty and valid_haircut.max() > 1 + tolerance else 1.0

    mask_haircut_100 = (df_main[haircut_col].sub(target_100_reset).abs() < tolerance)
    mask_below_threshold = (df_main[conc_calc_col] < threshold_limit)
    mask_final = mask_haircut_100 | mask_below_threshold

    df_main.loc[mask_final, conc_limit_col] = 0.0
    return df_main

def keterangan_uma(uma_date):
    if pd.notna(uma_date):
        if not isinstance(uma_date, datetime):
            try:
                uma_date = pd.to_datetime(str(uma_date))
            except Exception:
                return "Sesuai Metode Perhitungan"
        return f"Sesuai Haircut KPEI, mempertimbangkan pengumuman UMA dari BEI tanggal {uma_date.strftime('%d %b %Y')}"
    return "Sesuai Metode Perhitungan"

# ===============================================================
# FUNGSI UTAMA UNTUK CONCENTRATION LIMIT (CL)
# ===============================================================

def calculate_concentration_limit(df_cl_source: pd.DataFrame) -> pd.DataFrame:
    """Menjalankan seluruh logika perhitungan Concentration Limit."""
    
    df = df_cl_source.copy()
    
    if 'KODE EFEK' not in df.columns:
        df = df.rename(columns={df.columns[0]: 'KODE EFEK'})
        
    df['KODE EFEK'] = df['KODE EFEK'].astype(str).str.strip()
    
    # 1. Perhitungan limit marjin
    df['SAHAM MARJIN BARU?'] = df['SAHAM MARJIN BARU?'].astype(str).str.upper().str.strip()
    df['CONCENTRATION LIMIT KARENA SAHAM MARJIN BARU'] = np.where(
        df['SAHAM MARJIN BARU?'] == 'YA',
        df[COL_PERHITUNGAN] * 0.50,
        df[COL_PERHITUNGAN]
    )

    # 2. Hitung limit listed & FF
    df[COL_LISTED] = df.apply(calc_concentration_limit_listed, axis=1)
    df[COL_FF] = df.apply(calc_concentration_limit_ff, axis=1)

    # 3. & 4. TENTUKAN CONCENTRATION LIMIT USULAN RMCC (Ambil nilai minimum)
    limit_cols_for_min = [
        'CONCENTRATION LIMIT KARENA SAHAM MARJIN BARU', 
        COL_LISTED, COL_FF, COL_PERHITUNGAN
    ]

    df['MIN_CL_OPTION'] = df[limit_cols_for_min].fillna(np.inf).min(axis=1)
    
    # Tentukan pemicu nol
    mask_pemicu_nol = (
        (df['CONCENTRATION LIMIT KARENA SAHAM MARJIN BARU'].fillna(np.inf) < THRESHOLD_5M) |
        (df[COL_PERHITUNGAN].fillna(np.inf) < THRESHOLD_5M) |
        (df[COL_LISTED].fillna(np.inf) < THRESHOLD_5M) |
        (df[COL_FF].fillna(np.inf) < THRESHOLD_5M)
    )

    df[COL_RMCC] = np.where(
        mask_pemicu_nol,
        0.0, 
        df['MIN_CL_OPTION']
    )

    # 5. Override emiten khusus (Menerapkan CAP 10M/50M)
    mask_not_zero = (df[COL_RMCC] != 0.0)
    df.loc[mask_not_zero, COL_RMCC] = df.loc[mask_not_zero].apply(override_rmcc_limit, axis=1).round(0)

    # 6. Tambah kolom haircut usulan 
    df['HAIRCUT KPEI'] = pd.to_numeric(df['HAIRCUT KPEI'], errors='coerce').fillna(0)
    df['HAIRCUT PEI USULAN DIVISI'] = np.where(
        (df['UMA'].fillna('-') != '-') & pd.notna(df['UMA']),
        df['HAIRCUT KPEI'],
        df['HAIRCUT PEI']
    )

    # 7. Nolkan CL USULAN RMCC jika haircut awal 100% atau CL perhitungan < 5M
    df = reset_concentration_limit(df)

    # 7B. SET HAIRCUT JADI 100% KARENA CL=0
    HAIRCUT_COL_USULAN = 'HAIRCUT PEI USULAN DIVISI'
    mask_rmcc_nol = (df[COL_RMCC] == 0)
    df['TEMP_HAIRCUT_VAL_ASLI'] = pd.to_numeric(df[HAIRCUT_COL_USULAN], errors='coerce') 
    df.loc[mask_rmcc_nol, HAIRCUT_COL_USULAN] = 1.0 

    # 8. Keterangan Haircut & CONC LIMIT
    df['PERTIMBANGAN DIVISI (HAIRCUT)'] = df['UMA'].apply(keterangan_uma)
    df['PERTIMBANGAN DIVISI (CONC LIMIT)'] = 'Sesuai metode perhitungan' # Placeholder

    # 9. **(LOGIKA REVISI FINAL)** Definisikan Masker

    # --- Definisikan Semua Masker CL = 0 ---
    mask_emiten = df['KODE EFEK'].isin(KODE_EFEK_KHUSUS)
    mask_nol_final = (df[COL_RMCC] == 0) & (~mask_emiten)
    
    mask_lt5m_strict = (
        (df['CONCENTRATION LIMIT KARENA SAHAM MARJIN BARU'].fillna(np.inf) < THRESHOLD_5M) |
        (df[COL_PERHITUNGAN].fillna(np.inf) < THRESHOLD_5M) |
        (df[COL_LISTED].fillna(np.inf) < THRESHOLD_5M) |
        (df[COL_FF].fillna(np.inf) < THRESHOLD_5M) |
        (df[COL_RMCC].fillna(np.inf) < THRESHOLD_5M)
    ) & mask_nol_final 
                        
    mask_hc100_origin = ((df['TEMP_HAIRCUT_VAL_ASLI'].sub(1.0).abs() < TOLERANCE) |
                         (df['TEMP_HAIRCUT_VAL_ASLI'].sub(TARGET_100).abs() < TOLERANCE)) & \
                        mask_nol_final & \
                        (~mask_lt5m_strict)

    # --- Definisikan Masker CL Non-Nol (CL != 0) ---
    mask_non_nol = (df[COL_RMCC] != 0)
    
    # 1. Keterangan Emiten Khusus (Override Cap)
    df['CL_OVERRIDE_VAL'] = df['KODE EFEK'].map(OVERRIDE_MAPPING).fillna(np.inf)
    # True jika CL_RMCC sama persis dengan batas override (yang berarti override terpicu)
    mask_emiten_override = mask_non_nol & mask_emiten & \
                           (df[COL_RMCC].round(0) == df['CL_OVERRIDE_VAL'].round(0))

    mask_other_non_nol = mask_non_nol & (~mask_emiten_override)
    
    # Masker Pendukung Kriteria (Apakah emiten kena Listed/FF secara kriteria?)
    mask_kriteria_listed = pd.notna(df[COL_LISTED])
    mask_kriteria_ff = pd.notna(df[COL_FF])
    
    # Masker Sumber Nilai Minimum (HARUS BERDASARKAN KESAMAAN DENGAN CL_RMCC)
    
    # 2. CL karena Listed Saja (Berasal dari Minimum)
    mask_listed_saja = mask_other_non_nol & \
                       (df[COL_RMCC].round(0) == df[COL_LISTED].round(0))
    
    # 3. CL karena FF Saja (Berasal dari Minimum)
    # PENTING: Diberi pengecualian Listed Saja untuk menghindari konflik nilai yang sama
    mask_ff_saja = mask_other_non_nol & \
                   (df[COL_RMCC].round(0) == df[COL_FF].round(0)) & \
                   (~mask_listed_saja)
    
    # 4. CL karena Saham Marjin Baru (Berasal dari Minimum)
    # PENTING: Diberi pengecualian agar tidak bentrok jika nilainya sama persis dengan Listed/FF
    mask_marjin_baru = mask_other_non_nol & \
                       (df[COL_RMCC].round(0) == df['CONCENTRATION LIMIT KARENA SAHAM MARJIN BARU'].round(0)) & \
                       (df['SAHAM MARJIN BARU?'] == 'YA') & \
                       (~mask_listed_saja) & (~mask_ff_saja)

    # 5. LOGIKA UPGRADE KETERANGAN LISTED/FF GANDA
    # Jika CL_RMCC berasal dari Listed ATAU FF, DAN kriteria Listed & FF keduanya terpenuhi
    # (Ini akan menang atas Listed Saja atau FF Saja)
    mask_listed_ff_upgrade = (mask_listed_saja | mask_ff_saja) & \
                             mask_kriteria_listed & \
                             mask_kriteria_ff
    
    
    # 10. **(PENERAPAN FINAL)** Penerapan Keterangan (Prioritas Rendah ke Tinggi)

    # A. CL = 0: CL < 5M (Prioritas 1)
    df.loc[mask_lt5m_strict, 'PERTIMBANGAN DIVISI (CONC LIMIT)'] = 'Penyesuaian karena Batas Konsentrasi < Rp5 Miliar'
    df.loc[mask_lt5m_strict, 'PERTIMBANGAN DIVISI (HAIRCUT)'] = 'Penyesuaian karena Batas Konsentrasi 0'
    
    # B. CL = 0: HC 100% (Prioritas 2)
    df.loc[mask_hc100_origin, 'PERTIMBANGAN DIVISI (CONC LIMIT)'] = 'Penyesuaian karena Haircut PEI 100%'
    
    # C. CL = 0: Emiten Khusus (Prioritas 3)
    df.loc[mask_emiten & mask_nol_final, 'PERTIMBANGAN DIVISI (CONC LIMIT)'] = 'Penyesuaian karena profil emiten'
    
    # D. CL != 0: Emiten Khusus (Override Cap) (Prioritas 4)
    # Harus di bawah P5-P8, namun jika nilai CL_RMCC sama dengan Override, inilah keterangannya.
    df.loc[mask_emiten_override, 'PERTIMBANGAN DIVISI (CONC LIMIT)'] = 'Penyesuaian karena profil emiten'
    
    # E. CL != 0: SAHAM MARJIN BARU (Prioritas 5)
    df.loc[mask_marjin_baru, 'PERTIMBANGAN DIVISI (CONC LIMIT)'] = 'Penyesuaian karena saham baru masuk marjin'
    
    # F. CL != 0: Listed Saja (Prioritas 6)
    df.loc[mask_listed_saja, 'PERTIMBANGAN DIVISI (CONC LIMIT)'] = 'Penyesuaian karena melebihi 5% listed shares'
    
    # G. CL != 0: FF Saja (Prioritas 7)
    df.loc[mask_ff_saja, 'PERTIMBANGAN DIVISI (CONC LIMIT)'] = 'Penyesuaian karena melebihi 20% free float'
    
    # H. CL != 0: Listed/FF Ganda (Prioritas 8 - UPGRADE KETERANGAN)
    # Ini menimpa (F) atau (G) jika sumber minimumnya adalah Listed atau FF, tetapi secara kriteria, dia Ganda.
    df.loc[mask_listed_ff_upgrade, 'PERTIMBANGAN DIVISI (CONC LIMIT)'] = 'Penyesuaian karena melebihi 5% listed & 20% free float'
    
    # Cleanup kolom bantuan
    df = df.drop(columns=['MIN_CL_OPTION', 'TEMP_HAIRCUT_VAL_ASLI', 'CL_OVERRIDE_VAL'], errors='ignore')

    return df

# ============================
# ANTARMUKA CL
# ============================

def main():
    st.title("🛡️ Concentration Limit (CL) & Haircut Calculation")
    
    current_month_name = datetime.now().strftime('%B').lower()
    example_filename = f'HCCL_{current_month_name}.xlsx'
    
    st.markdown(f"Unggah file sumber data Concentration Limit (misal: `{example_filename}` atau sejenisnya) untuk menjalankan perhitungan CL.")
    
    required_file_cl = f'File Sumber Concentration Limit (misal: {example_filename})'
    
    uploaded_file_cl = st.file_uploader(f"Unggah {required_file_cl}", type=['xlsx'], key='cl_source')
    
    if uploaded_file_cl is not None:
        if st.button("Jalankan Perhitungan CL", type="primary"):
            try:
                df_cl_source = pd.read_excel(uploaded_file_cl, engine='openpyxl')
                
                with st.spinner('Menghitung Concentration Limit...'):
                    df_cl_hasil = calculate_concentration_limit(df_cl_source)
                
                st.success("✅ Perhitungan Concentration Limit selesai. Siap diunduh!")
                st.subheader("Hasil Concentration Limit (Tabel)")
                st.dataframe(df_cl_hasil) 
                
                output_buffer_cl = BytesIO()
                df_cl_hasil.to_excel(output_buffer_cl, index=False)
                output_buffer_cl.seek(0)
                
                month_name_lower_output = datetime.now().strftime('%B').lower()
                dynamic_filename_output = f'clhc_{month_name_lower_output}.xlsx' 
                
                st.download_button(
                    label="⬇️ Unduh Hasil Concentration Limit",
                    data=output_buffer_cl,
                    file_name=dynamic_filename_output, 
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )

            except Exception as e:
                st.error(f"❌ Gagal dalam perhitungan CL. Pastikan format file benar. Error: {e}")

if __name__ == '__main__':
    main()



